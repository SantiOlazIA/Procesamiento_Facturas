"""Check if column V numeration and formatting worked."""
import openpyxl

wb = openpyxl.load_workbook(
    r"C:\Users\Tuchi\MiEstudioIA\FCI Santander\Output\FCI_Procesado_FCI_SUPER_AHORRO_CL_H_PJ.xlsx"
)
ws = wb.active

print("=== Check headers ===")
print(f"  V4='{ws['V4'].value}', W4='{ws['W4'].value}'")

print("\n=== Check Cuadro rows 205-220 ===")
for r in range(205, 222):
    v = ws[f'V{r}'].value
    w = ws[f'W{r}'].value
    x = ws[f'X{r}'].value
    z = ws[f'Z{r}'].value
    if v or w or x or z:
        print(f"  Row {r}: V={v} W={w} X={x} Z={z}")

# Check borders on Cuadro rows
print("\n=== Check borders on rows 5 and 215 ===")
for r in [5, 215]:
    borders_info = []
    for c in ['V', 'W', 'Z', 'AG']:
        border = ws[f'{c}{r}'].border
        has_top = bool(border.top.style) if border.top else False
        has_bottom = bool(border.bottom.style) if border.bottom else False
        borders_info.append(f"{c}:T={has_top},B={has_bottom}")
    print(f"  Row {r} borders: " + " | ".join(borders_info))
