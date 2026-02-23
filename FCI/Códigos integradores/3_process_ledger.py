import openpyxl
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from copy import copy
import json
import datetime
import os
import sys

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(SCRIPT_DIR)
from config import OUTPUT_EXCEL_PATH, TEMPLATE_EXCEL_PATH, JSON_DATA_PATH, VIBRANT_COLORS, FONT_NAME, FONT_SIZE, BORDER_STYLE, DATE_FORMAT, log

medium = Side(border_style='medium', color="000000")

def apply_sub_borders(ws, r_a):
    """Apply REF-accurate borders for a Subscription asiento (4 rows: r_a to r_a+3)."""
    ws[f'B{r_a}'].border = Border(top=medium, left=medium, right=medium)
    ws[f'C{r_a}'].border = Border(top=medium)
    ws[f'D{r_a}'].border = Border(top=medium)
    ws[f'E{r_a}'].border = Border(top=medium)
    ws[f'F{r_a}'].border = Border(top=medium, bottom=medium, left=medium, right=medium)
    ws[f'G{r_a}'].border = Border(top=medium, left=medium, right=medium)
    ws[f'H{r_a}'].border = Border(top=medium, right=medium)
    for r in range(r_a + 1, r_a + 3):
        ws[f'B{r}'].border = Border(left=medium, right=medium)
        ws[f'G{r}'].border = Border(left=medium, right=medium)
        ws[f'H{r}'].border = Border(right=medium)
    foot = r_a + 3
    ws[f'B{foot}'].border = Border(bottom=medium, left=medium, right=medium)
    ws[f'C{foot}'].border = Border(bottom=medium)
    ws[f'D{foot}'].border = Border(bottom=medium)
    ws[f'E{foot}'].border = Border(bottom=medium)
    ws[f'F{foot}'].border = Border(bottom=medium)
    ws[f'G{foot}'].border = Border(bottom=medium, left=medium, right=medium)
    ws[f'H{foot}'].border = Border(bottom=medium, right=medium)

def apply_rescue_borders(ws, r_a):
    """Apply REF-accurate borders for a Rescue asiento (5 rows: r_a to r_a+4)."""
    ws[f'B{r_a}'].border = Border(top=medium, left=medium, right=medium)
    ws[f'C{r_a}'].border = Border(top=medium)
    ws[f'D{r_a}'].border = Border(top=medium)
    ws[f'E{r_a}'].border = Border(top=medium)
    ws[f'F{r_a}'].border = Border(top=medium, bottom=medium, left=medium, right=medium)
    ws[f'G{r_a}'].border = Border(top=medium, left=medium, right=medium)
    ws[f'H{r_a}'].border = Border(top=medium, right=medium)
    for r in range(r_a + 1, r_a + 4):
        ws[f'B{r}'].border = Border(left=medium, right=medium)
        ws[f'G{r}'].border = Border(left=medium, right=medium)
        ws[f'H{r}'].border = Border(right=medium)
    foot = r_a + 4
    ws[f'B{foot}'].border = Border(bottom=medium, left=medium, right=medium)
    ws[f'C{foot}'].border = Border(bottom=medium)
    ws[f'D{foot}'].border = Border(bottom=medium)
    ws[f'E{foot}'].border = Border(bottom=medium)
    ws[f'F{foot}'].border = Border(bottom=medium)
    ws[f'G{foot}'].border = Border(bottom=medium, left=medium, right=medium)
    ws[f'H{foot}'].border = Border(bottom=medium, right=medium)

def process_ledger():
    log.info("=" * 50)
    log.info("PASO 3: Procesamiento del Libro Mayor")
    log.info("=" * 50)
    
    if os.path.exists(OUTPUT_EXCEL_PATH):
        log.info(f"Cargando libro existente")
        wb = openpyxl.load_workbook(OUTPUT_EXCEL_PATH)
    else:
        log.info(f"Creando libro nuevo desde template")
        if not os.path.exists(TEMPLATE_EXCEL_PATH):
            log.error(f"Template no encontrado: {TEMPLATE_EXCEL_PATH}")
            return False
        wb = openpyxl.load_workbook(TEMPLATE_EXCEL_PATH)
        
    ws = wb.active 
    
    # Clear stale Traducción rows (O-U)
    log.debug("Limpiando filas de Traduccion antiguas...")
    for r in range(4, 50000):
        for c in ['O','P','Q','R','S','T','U']:
            if ws[f'{c}{r}'].value is not None:
                ws[f'{c}{r}'].value = None
    
    last_row_cuadro = 5
    for r in range(5, 50000):
        if ws[f'W{r}'].value: last_row_cuadro = r
    
    last_date_in_excel = None
    if last_row_cuadro >= 5:
        val = ws[f'W{last_row_cuadro}'].value
        if isinstance(val, datetime.datetime): last_date_in_excel = val
        elif isinstance(val, str):
             try: last_date_in_excel = datetime.datetime.strptime(val, "%d/%m/%Y")
             except: pass

    subscriptions = []
    active_sub_index = 0
    suscrip_counter = 1
    rescate_counter = 1
    
    for r in range(5, last_row_cuadro + 1):
        concept = str(ws[f'X{r}'].value)
        if "SUSCRIPCIÓN" in concept or "SUSCRIPCION" in concept:
            try: sid = int(concept.split(' ')[1]); suscrip_counter = max(suscrip_counter, sid + 1)
            except: pass
            z_val = ws[f'Z{r}'].value; cp = float(z_val) if z_val else 0.0
            fill = ws[f'Z{r}'].fill; price_cell = f'AA{r}'
            subscriptions.append({'id': sid, 'remaining_cp': cp, 'fill': fill, 'price_cell': price_cell, 'row': r})
        elif "RESCATE" in concept or "VENTA" in concept:
            if "RESCATE" in concept:
                try: rid = int(concept.split(' ')[1]); rescate_counter = max(rescate_counter, rid + 1)
                except: pass
            z_val = ws[f'Z{r}'].value
            if z_val:
                consumed = abs(float(z_val))
                while consumed > 0.001 and active_sub_index < len(subscriptions):
                    curr = subscriptions[active_sub_index]
                    if curr['remaining_cp'] > consumed:
                        curr['remaining_cp'] -= consumed; consumed = 0
                    else:
                        consumed -= curr['remaining_cp']; curr['remaining_cp'] = 0; active_sub_index += 1
                        
    current_color_idx = max(0, suscrip_counter - 2)
    log.debug(f"Estado FIFO: {len(subscriptions)} suscripciones cargadas, indice activo={active_sub_index}")
    
    if not os.path.exists(JSON_DATA_PATH):
        log.error(f"Archivo JSON no encontrado: {JSON_DATA_PATH}")
        return False
    with open(JSON_DATA_PATH, 'r', encoding='utf-8') as f: new_movs = json.load(f)
    
    to_process = []
    to_process = []
    # Check for Start Date (Oct 1 2024)
    start_date_filter = datetime.datetime(2024, 10, 1)

    if last_date_in_excel and last_date_in_excel > start_date_filter:
        log.info(f"Filtrando >= {last_date_in_excel.strftime('%d/%m/%Y')}")
        filter_date = last_date_in_excel
    else:
        # Default start of period
        log.info(f"Iniciando periodo desde {start_date_filter.strftime('%d/%m/%Y')}")
        filter_date = start_date_filter

    for m in new_movs:
        try:
            d = datetime.datetime.strptime(m['date'], "%d/%m/%Y")
            if d >= filter_date: 
                to_process.append(m)
        except: pass
    
    log.info(f"Movimientos a procesar: {len(to_process)}")
    if not to_process:
        log.info("Sin movimientos nuevos")
        return True

    current_cuadro_row = last_row_cuadro + 1
    last_row_asiento = 4 
    for r in range(4, 50000):
        if ws[f'B{r}'].value or ws[f'C{r}'].value or ws[f'D{r}'].value: last_row_asiento = r
    current_asiento_row = last_row_asiento + 1
    
    asiento_blocks = []
    
    # Scan existing asientos
    for r in range(4, current_asiento_row):
        d_val = ws[f'D{r}'].value
        if d_val and str(d_val).startswith('='):
            i_val = ws[f'I{r}'].value
            if i_val:
                asiento_blocks.append((r, 'RESCUE', 5))
            else:
                asiento_blocks.append((r, 'SUB', 4))
    
    sub_count = 0
    resc_count = 0
    
    for mov in to_process:
        m_date = datetime.datetime.strptime(mov['date'], "%d/%m/%Y")
        cp = mov['cp']
        amt = mov['amount']
        
        if mov['type'] == 'COMPRA':
            r_c = current_cuadro_row
            ws[f'W{r_c}'] = m_date; ws[f'W{r_c}'].number_format = 'DD/MM/YYYY'
            ws[f'X{r_c}'] = f"SUSCRIPCIÓN {suscrip_counter}"
            ws[f'Y{r_c}'] = 1
            ws[f'Z{r_c}'] = cp
            ws[f'AB{r_c}'] = amt
            ws[f'AA{r_c}'] = f'=AB{r_c}/Z{r_c}'
            
            if r_c == 5: ws[f'AC{r_c}'] = f'=Z{r_c}'; ws[f'AE{r_c}'] = f'=AB{r_c}'
            else:
                 ws[f'AC{r_c}'] = f'=AC{r_c-1}+Z{r_c}'
                 ws[f'AE{r_c}'] = f'=AE{r_c-1}+AB{r_c}+AD{r_c}'
            ws[f'AD{r_c}'] = 0
            ws[f'AF{r_c}'] = f'=IF(+AB{r_c}>1,+AB{r_c},-AB{r_c})/AA{r_c}'
            ws[f'AG{r_c}'] = f'=+AF{r_c}-IF(Y{r_c}=1,+Z{r_c},-Z{r_c})'
            
            c_hex = VIBRANT_COLORS[current_color_idx % len(VIBRANT_COLORS)]
            fill = PatternFill(start_color=c_hex, end_color=c_hex, fill_type='solid')
            ws[f'Z{r_c}'].fill = fill
            current_color_idx += 1
            subscriptions.append({'id': suscrip_counter, 'remaining_cp': cp, 'fill': fill, 'price_cell': f'AA{r_c}', 'row': r_c})
            suscrip_counter += 1
            sub_count += 1
            
            r_a = current_asiento_row
            ws[f'D{r_a}'] = f'=+W{r_c}'; ws[f'D{r_a}'].number_format = 'DD/MM/YYYY'
            ws[f'N{r_a}'] = f'=+X{r_c}'
            ws[f'B{r_a+1}'] = '10201-1'; ws[f'C{r_a+1}'] = 'Fondo Común de Inv BBVA'; ws[f'G{r_a+1}'] = f'=ROUND(AB{r_c}, 2)' 
            ws[f'B{r_a+2}'] = '10101'; ws[f'D{r_a+2}'] = 'Caja Banco'; ws[f'H{r_a+2}'] = f'=ROUND(AB{r_c}, 2)' 
            ws[f'C{r_a+3}'] = 'Suscripción cuotas partes FCI BBVA'
            
            apply_sub_borders(ws, r_a)
            asiento_blocks.append((r_a, 'SUB', 4))
            current_cuadro_row += 1
            current_asiento_row += 4
            
        else:  # VENTA
            needed = cp; quote_val = amt / cp if cp > 0 else 0
            while needed > 0.0001:
                if active_sub_index >= len(subscriptions):
                    log.warning(f"  FIFO agotado procesando VENTA del {mov['date']} (faltan {needed:.2f} CPs)")
                    break
                sub = subscriptions[active_sub_index]
                consume = 0
                if sub['remaining_cp'] >= needed: consume = needed; sub['remaining_cp'] -= needed; needed = 0
                else: consume = sub['remaining_cp']; needed -= consume; sub['remaining_cp'] = 0; active_sub_index += 1
                
                r_c = current_cuadro_row
                ws[f'W{r_c}'] = m_date; ws[f'W{r_c}'].number_format = 'DD/MM/YYYY'
                ws[f'X{r_c}'] = f"RESCATE {rescate_counter}"
                ws[f'Y{r_c}'] = 2
                ws[f'Z{r_c}'] = -consume
                ws[f'Z{r_c}'].fill = copy(sub['fill'])
                ws[f'AB{r_c}'] = -consume * quote_val 
                ws[f'AA{r_c}'] = f'=AB{r_c}/Z{r_c}' 
                
                if r_c == 5: ws[f'AC{r_c}'] = f'=Z{r_c}'; ws[f'AE{r_c}'] = f'=AB{r_c}'
                else:
                    ws[f'AC{r_c}'] = f'=AC{r_c-1}+Z{r_c}'
                    ws[f'AE{r_c}'] = f'=AE{r_c-1}+AB{r_c}+AD{r_c}'
                ws[f'AD{r_c}'] = f'=+H{current_asiento_row + 3}'
                ws[f'AF{r_c}'] = f'=IF(+AB{r_c}>1,+AB{r_c},-AB{r_c})/AA{r_c}'
                ws[f'AG{r_c}'] = f'=+AF{r_c}-IF(Y{r_c}=1,+Z{r_c},-Z{r_c})'
                
                r_a = current_asiento_row
                ws[f'D{r_a}'] = f'=+W{r_c}'; ws[f'D{r_a}'].number_format = 'DD/MM/YYYY'
                ws[f'N{r_a}'] = f'=+X{r_c}'
                ws[f'I{r_a}'] = f'=AA{r_c}'; ws[f'L{r_a}'] = f'=H{r_a+3}'; ws[f'K{r_a}'] = f'=L{r_a}/(-Z{r_c})'; ws[f'J{r_a}'] = f'=K{r_a}-I{r_a}'; ws[f'M{r_a}'] = f'=L{r_a}-H{r_a+3}'
                
                ws[f'B{r_a+1}'] = '10101'; ws[f'C{r_a+1}'] = 'Caja Banco'; ws[f'G{r_a+1}'] = f'=ROUND(-AB{r_c}, 2)' 
                ws[f'B{r_a+2}'] = '10201-1'; ws[f'D{r_a+2}'] = 'Fondo Común de Inv BBVA'; ws[f'H{r_a+2}'] = f'=ROUND(-Z{r_c}*{sub["price_cell"]}, 2)'
                ws[f'B{r_a+3}'] = '40203-1'; ws[f'D{r_a+3}'] = 'Rentas FCI BBVA'; ws[f'H{r_a+3}'] = f'=ROUND(G{r_a+1}-H{r_a+2}, 2)'
                ws[f'C{r_a+4}'] = 'Rescate cuotas partes de FCI BBVA'
                
                apply_rescue_borders(ws, r_a)
                asiento_blocks.append((r_a, 'RESCUE', 5))
                current_cuadro_row += 1
                current_asiento_row += 5
                resc_count += 1
            
            rescate_counter += 1
    
    # Generate Traducción rows (O-U) for ALL asiento blocks
    log.info(f"Generando Traduccion para {len(asiento_blocks)} bloques de asiento...")
    for header_row, block_type, block_size in asiento_blocks:
        if block_type == 'SUB':
            desc_row = header_row + 3
            for t in range(2):
                tr = header_row + t
                br = header_row + 1 + t
                ws[f'O{tr}'] = f'=+D{header_row}'
                ws[f'P{tr}'] = f'=MID(B{br},1,5)'
                ws[f'Q{tr}'] = f'=IF(MID(B{br},7,2)<>"",MID(B{br},7,2),0)'
                ws[f'R{tr}'] = 0
                ws[f'S{tr}'] = f'=ROUND(G{br},2)'
                ws[f'T{tr}'] = f'=ROUND(H{br},2)'
                ws[f'U{tr}'] = f'=+C{desc_row}'
        else:
            desc_row = header_row + 4
            for t in range(3):
                tr = header_row + t
                br = header_row + 1 + t
                ws[f'O{tr}'] = f'=+D{header_row}'
                ws[f'P{tr}'] = f'=MID(B{br},1,5)'
                ws[f'Q{tr}'] = f'=IF(MID(B{br},7,2)<>"",MID(B{br},7,2),0)'
                ws[f'R{tr}'] = 0
                ws[f'S{tr}'] = f'=ROUND(G{br},2)'
                ws[f'T{tr}'] = f'=ROUND(H{br},2)'
                ws[f'U{tr}'] = f'=+C{desc_row}'
            
    # ===== FORMATTING PASS =====
    log.info("Aplicando formato profesional al cuadro...")
    
    bold_font = Font(name='Arial', size=10, bold=True)
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(name='Arial', size=10, bold=True, color='FFFFFF')
    thin = Side(border_style='thin', color='000000')
    header_border = Border(top=thin, bottom=thin, left=thin, right=thin)
    
    # Number formats
    FMT_MONEY = '#,##0.00'       # 685.714.581,66
    FMT_CPS = '#,##0.00'         # 5.909.501,57
    FMT_PRICE = '#,##0.0000000000'  # 116.0359420397
    FMT_DATE = 'DD/MM/YYYY'
    
    # Header styling (row 4: W-AG)
    headers = {
        'W': 'Fecha', 'X': 'Concepto', 'Y': 'Código',
        'Z': 'Cuotas parte', 'AA': 'Cotización', 'AB': 'Importe Cta Cte',
        'AC': 'Saldo en CP', 'AD': 'Renta rescate', 'AE': 'Saldo en $',
        'AF': 'Prueba', 'AG': 'Prueba'
    }
    for col, label in headers.items():
        cell = ws[f'{col}4']
        cell.value = label
        cell.font = header_font
        cell.fill = header_fill
        cell.border = header_border
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
    
    # Column widths (AutoFit approximation)
    col_widths = {
        'W': 12, 'X': 18, 'Y': 8, 'Z': 16,
        'AA': 16, 'AB': 18, 'AC': 16, 'AD': 14,
        'AE': 18, 'AF': 14, 'AG': 14
    }
    for col, width in col_widths.items():
        ws.column_dimensions[col].width = width
    
    # Apply number formats to all data rows
    last_data_row = current_cuadro_row - 1
    for r in range(5, last_data_row + 1):
        ws[f'W{r}'].number_format = FMT_DATE
        ws[f'Z{r}'].number_format = FMT_CPS
        ws[f'AA{r}'].number_format = FMT_PRICE
        ws[f'AB{r}'].number_format = FMT_MONEY
        ws[f'AC{r}'].number_format = FMT_CPS
        ws[f'AD{r}'].number_format = FMT_MONEY
        ws[f'AE{r}'].number_format = FMT_MONEY
        ws[f'AF{r}'].number_format = FMT_CPS
        ws[f'AG{r}'].number_format = FMT_CPS
    
    # Also format asiento G/H columns (Debe/Haber) 
    for r in range(4, current_asiento_row):
        if ws[f'G{r}'].value is not None:
            ws[f'G{r}'].number_format = FMT_MONEY
        if ws[f'H{r}'].value is not None:
            ws[f'H{r}'].number_format = FMT_MONEY
    
    wb.save(OUTPUT_EXCEL_PATH)
    
    # Summary
    log.info(f"\nResumen de procesamiento:")
    log.info(f"  Suscripciones generadas: {sub_count}")
    log.info(f"  Rescates generados: {resc_count}")
    log.info(f"  Filas cuadro: {current_cuadro_row - 5}")
    log.info(f"  Bloques asiento: {len(asiento_blocks)}")
    log.info(f"  Guardado en: {OUTPUT_EXCEL_PATH}")
    return True

if __name__ == "__main__":
    process_ledger()
