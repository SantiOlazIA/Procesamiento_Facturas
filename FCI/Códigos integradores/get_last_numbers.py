import openpyxl
import os
import re

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, 'FCI_Procesado_Anual.xlsx')

def get_last_numbers():
    if not os.path.exists(EXCEL_PATH):
        print("Archivo no encontrado")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    
    max_sub = 0
    max_res = 0
    
    last_row = ws.max_row
    
    for r in range(5, last_row + 1):
        concept = ws[f'X{r}'].value
        if concept and isinstance(concept, str):
            concept = concept.upper()
            # Check for SUSCRIPCIÓN X
            m_sub = re.search(r'SUSCRIPCI[ÓO]N\s+(\d+)', concept)
            if m_sub:
                val = int(m_sub.group(1))
                if val > max_sub: max_sub = val
            
            # Check for RESCATE X
            m_res = re.search(r'RESCATE\s+(\d+)', concept)
            if m_res:
                val = int(m_res.group(1))
                if val > max_res: max_res = val

    print(f"Última Suscripción: {max_sub}")
    print(f"Último Rescate:     {max_res}")

if __name__ == "__main__":
    get_last_numbers()
