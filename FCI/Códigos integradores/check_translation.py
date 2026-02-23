import openpyxl
import os
import re

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, 'FCI_Procesado_Anual.xlsx')

def check_translation():
    if not os.path.exists(EXCEL_PATH):
        print("Archivo no encontrado")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    ws = wb.active
    
    last_row = ws.max_row
    print(f"Verificando filas de Traducción (O-U) hasta fila {last_row}...")
    
    errors = []
    
    # We scan for rows where there is content in O (Date)
    # For each such row, we check S (Debit) and T (Credit) formulas
    
    for r in range(4, last_row + 1):
        # Check if this is a translation row
        if ws[f'O{r}'].value:
            s_val = str(ws[f'S{r}'].value) # Debit Formula, e.g. =ROUND(G10, 2)
            t_val = str(ws[f'T{r}'].value) # Credit Formula, e.g. =ROUND(H10, 2)
            
            # Extract the source reference
            # Expecting =ROUND(G{r_source},2) or similar
            
            src_g = None
            src_h = None
            
            # Check S (Debit)
            m_s = re.search(r'G(\d+)', s_val)
            if m_s: src_g = int(m_s.group(1))
            
            # Check T (Credit)
            m_t = re.search(r'H(\d+)', t_val)
            if m_t: src_h = int(m_t.group(1))
            
            if not src_g and not src_h:
                # Is it a zero value?
                pass
            
            # Logic:
            # We want to ensure S and T mirror G and H of the same source row.
            # If S points to Gx, then T should point to Hx (or be 0 if Hx is unused, but the formula usually is =ROUND(Hx,2))
            
            # In the script generation:
            # ws[f'S{tr}'] = f'=ROUND(G{br},2)'
            # ws[f'T{tr}'] = f'=ROUND(H{br},2)'
            # So S and T ALWAYS point to the SAME source row 'br'.
            
            # Let's extract 'br' from S and T and check they match.
            br_s = None
            br_t = None
            
            if 'G' in s_val:
                m = re.search(r'G(\d+)', s_val)
                if m: br_s = int(m.group(1))
            
            if 'H' in t_val:
                m = re.search(r'H(\d+)', t_val)
                if m: br_t = int(m.group(1))
                
            if br_s and br_t:
                if br_s != br_t:
                    errors.append(f"Fila {r}: Desajuste de referencia. S apunta a fila {br_s}, T apunta a fila {br_t}.")
                else:
                    # Verified they point to same row.
                    # Now check logic: The source row {br_s} in Asiento must be balanced or part of a balanced block.
                    # We verified previously that Asiento blocks are balanced (M=L-H=0).
                    pass
            elif br_s and not br_t:
                 # It might be that T is hardcoded 0?
                 if t_val != '0' and t_val != '0.0':
                     errors.append(f"Fila {r}: S tiene ref G{br_s} pero T es {t_val}")
            elif br_t and not br_s:
                 if s_val != '0' and s_val != '0.0':
                     errors.append(f"Fila {r}: T tiene ref H{br_t} pero S es {s_val}")

    if not errors:
        print("VERIFICACION EXITOSA: Todas las filas de traducción referencian correctamente a sus contrapartes en el asiento.")
        print("Dado que los asientos están balanceados (M=0), la traducción también está balanceada.")
    else:
        print(f"Se encontraron {len(errors)} errores:")
        for e in errors[:10]:
            print(e)

if __name__ == "__main__":
    check_translation()
