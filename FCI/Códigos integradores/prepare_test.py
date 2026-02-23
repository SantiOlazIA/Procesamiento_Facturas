"""
Prepare Test Environment — Mejora 2: Auto-detect initial subscription from PDFs.

Reads the first COMPRA movement from movements.json to seed the initial
subscription automatically, eliminating manual editing.
"""
import openpyxl
from openpyxl.styles import PatternFill, Border, Side
import os
import sys
import shutil
import datetime
import json

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(SCRIPT_DIR)
from config import (INPUT_PDF_DIR, TEMPLATE_EXCEL_PATH, OUTPUT_EXCEL_PATH,
                    JSON_DATA_PATH, VIBRANT_COLORS, log)

TEMPLATE_SRC = os.path.join(os.path.dirname(SCRIPT_DIR), 'Input', 'Formato FCI v2.xlsx')
medium = Side(border_style='medium', color="000000")



from pypdf import PdfReader
import re

def get_sept_2024_closing_balance():
    """Reads the Sept 2024 PDF to find the closing balance (Saldo Final)."""
    fname = "FONDO COMUN DE INVERSION 24 09.pdf"
    fpath = os.path.join(INPUT_PDF_DIR, fname)
    
    if not os.path.exists(fpath):
        log.warning(f"No se encontro {fname} para extraer saldo inicial")
        return None

    try:
        reader = PdfReader(fpath)
        text = ""
        for page in reader.pages:
            text += page.extract_text() + "\n"
        
        # Regex to find: FBA RENPEB ... 5909.501,57 116,03594200 685.714.581,66
        # Matches the 'SALDOS DISPONIBLES' section line often starting with concept
        # We look for the specific values structure at end of line
        pattern = r'FBA RENPEB.*?\s+([\d\.,]+)\s+([\d\.,]+)\s+([\d\.,]+)$'
        
        for line in text.split('\n'):
            line = line.strip()
            # Look for line ending with values (CP, Price, Total)
            # The PDF line inspected was:
            # FBA RENPEB 22886.174,02 -16976.672,45 5909.501,57 116,03594200 685.714.581,66
            if "FBA RENPEB" in line and "116,03594200" in line: # Use known price to identify line safely if possible, or just regex
                 match = re.search(r'FBA RENPEB.*?\s+([\d\.]+,\d{2})\s+([\d\.,]+)\s+([\d\.]+,\d{2})$', line)
                 if match:
                     cp_str = match.group(1).replace('.', '').replace(',', '.')
                     amt_str = match.group(3).replace('.', '').replace(',', '.')
                     return float(cp_str), float(amt_str)
                     
        log.warning("No se pudo extraer el saldo final del texto del PDF")
        return None
        
    except Exception as e:
        log.error(f"Error leyendo {fname}: {e}")
        return None


def prepare():
    """
    Prepare clean template with Initial Subscription from Sept 2024 PDF.
    """
    log.info("=" * 50)
    log.info("PREPARACION DEL ENTORNO")
    log.info("=" * 50)
    
    if os.path.exists(OUTPUT_EXCEL_PATH):
        os.remove(OUTPUT_EXCEL_PATH)
        log.info("Eliminado output anterior")
    
    if not os.path.exists(TEMPLATE_SRC):
        log.error(f"Template fuente no encontrado: {TEMPLATE_SRC}")
        return False

    shutil.copy2(TEMPLATE_SRC, TEMPLATE_EXCEL_PATH)
    wb = openpyxl.load_workbook(TEMPLATE_EXCEL_PATH)
    ws = wb.active

    # Clear cuadro data
    for r in range(5, 50000):
        for c in ['W','X','Y','Z','AA','AB','AC','AD','AE','AF','AG']:
            if ws[f'{c}{r}'].value is not None:
                ws[f'{c}{r}'].value = None
                ws[f'{c}{r}'].fill = PatternFill(fill_type=None)

    # Clear asiento data AND ALL BORDERS
    for r in range(4, 50000):
        for c in ['B','C','D','E','F','G','H','I','J','K','L','M','N','O','P','Q','R','S','T','U']:
            if ws[f'{c}{r}'].value is not None:
                ws[f'{c}{r}'].value = None
        for c in ['B','C','D','E','F','G','H']:
            ws[f'{c}{r}'].border = Border()

    # Get Initial Balance from Sept 2024 PDF
    initial_data = get_sept_2024_closing_balance()
    
    if initial_data:
        initial_cp, initial_amount = initial_data
        initial_date = datetime.datetime(2024, 10, 1) # Start of Period
        log.info(f"  Saldo Inicial (desde PDF Sept 24):")
    else:
        # Fallback (Manual override or error) - hardcoded from inspection if PDF read fails
        log.warning("  Usando valores hardcodeados de respaldo (Inspeccion previa)")
        initial_date = datetime.datetime(2024, 10, 1)
        initial_cp = 5909501.57
        initial_amount = 685714581.66
    
    initial_price = initial_amount / initial_cp if initial_cp > 0 else 0

    log.info(f"    Fecha: {initial_date.strftime('%d/%m/%Y')}")
    log.info(f"    CPs: {initial_cp:,.2f}")
    log.info(f"    Monto: ${initial_amount:,.2f}")
    log.info(f"    Precio: ${initial_price:,.10f}")

    # Inject Initial Subscription
    ws['W5'] = initial_date
    ws['W5'].number_format = 'DD/MM/YYYY'
    ws['X5'] = 'SUSCRIPCIÓN 1'
    ws['Y5'] = 1
    ws['Z5'] = initial_cp
    ws['AA5'] = initial_price
    ws['AB5'] = initial_amount
    ws['AC5'] = '=Z5'
    ws['AD5'] = None
    ws['AE5'] = '=AB5'
    ws['AF5'] = '=IF(+AB5>1,+AB5,-AB5)/AA5'
    ws['AG5'] = '=+AF5-IF(Y5=1,+Z5,-Z5)'

    fill = PatternFill(start_color=VIBRANT_COLORS[0], end_color=VIBRANT_COLORS[0], fill_type='solid')
    ws['Z5'].fill = fill

    # Initial Asiento -- REMOVED AS PER USER REQUEST
    # The subscription remains in the Table (Row 5) but no seat is generated.
    # ws['D4'] = '=+W5'; ws['D4'].number_format = 'DD/MM/YYYY'
    # ws['N4'] = '=+X5'
    # ws['B5'] = '10201-1'; ws['C5'] = 'Fondo Común de Inv BBVA'; ws['G5'] = '=AB5'
    # ws['B6'] = '10101'; ws['D6'] = 'Caja Banco'; ws['H6'] = '=+G5'
    # ws['C7'] = 'Suscripción cuotas partes FCI BBVA'

    # Borders (REF pattern) -- REMOVED
    # ws['B4'].border = Border(top=medium, left=medium, right=medium)
    # ws['C4'].border = Border(top=medium)
    # ws['D4'].border = Border(top=medium)
    # ws['E4'].border = Border(top=medium)
    # ws['F4'].border = Border(top=medium, bottom=medium, left=medium, right=medium)
    # ws['G4'].border = Border(top=medium, left=medium, right=medium)
    # ws['H4'].border = Border(top=medium, right=medium)
    # for r in [5, 6]:
    #     ws[f'B{r}'].border = Border(left=medium, right=medium)
    #     ws[f'G{r}'].border = Border(left=medium, right=medium)
    #     ws[f'H{r}'].border = Border(right=medium)
    # ws['B7'].border = Border(bottom=medium, left=medium, right=medium)
    # ws['C7'].border = Border(bottom=medium)
    # ws['D7'].border = Border(bottom=medium)
    # ws['E7'].border = Border(bottom=medium)
    # ws['F7'].border = Border(bottom=medium)
    # ws['G7'].border = Border(bottom=medium, left=medium, right=medium)
    # ws['H7'].border = Border(bottom=medium, right=medium)

    # Traducción -- REMOVED
    # ws['O4'] = '=+D4'; ws['P4'] = '=+MID(B5,1,5)'; ws['Q4'] = '=IF(MID(B5,7,2)<>"",MID(B5,7,2),0)'; ws['R4'] = 0; ws['S4'] = '=+ROUND(G5,2)'; ws['T4'] = '=+ROUND(H5,2)'; ws['U4'] = '=+C7'
    # ws['O5'] = '=+D4'; ws['P5'] = '=+MID(B6,1,5)'; ws['Q5'] = '=IF(MID(B6,7,2)<>"",MID(B6,7,2),0)'; ws['R5'] = 0; ws['S5'] = '=+ROUND(G6,2)'; ws['T5'] = '=+ROUND(H6,2)'; ws['U5'] = '=+C7'

    wb.save(TEMPLATE_EXCEL_PATH)
    log.info("Template preparado correctamente")
    return True


if __name__ == "__main__":
    prepare()
