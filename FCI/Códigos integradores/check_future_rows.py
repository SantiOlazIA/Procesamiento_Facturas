import openpyxl
import os
import datetime

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, 'FCI_Procesado_Anual.xlsx')

def check_future_rows():
    if not os.path.exists(EXCEL_PATH):
        print("Archivo no encontrado")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=True)
    ws = wb.active
    
    found_nov = False
    found_dec = False
    
    last_row = ws.max_row
    print(f"Escaneando hasta fila {last_row}...")
    
    for r in range(5, last_row + 1):
        date_val = ws[f'W{r}'].value
        if isinstance(date_val, datetime.datetime):
            if date_val.year == 2025:
                if date_val.month == 11:
                    found_nov = True
                    concept = ws[f'X{r}'].value
                    amt = ws[f'AB{r}'].value
                    print(f"  Encontrado Nov 2025 (Fila {r}): {date_val.strftime('%d/%m/%Y')} - {concept} - ${amt:,.2f}")
                elif date_val.month == 12:
                    found_dec = True
                    concept = ws[f'X{r}'].value
                    amt = ws[f'AB{r}'].value
                    print(f"  Encontrado Dic 2025 (Fila {r}): {date_val.strftime('%d/%m/%Y')} - {concept} - ${amt:,.2f}")

    if found_nov and found_dec:
        print("\nVERIFICACION EXITOSA: Se encontraron movimientos de Noviembre y Diciembre 2025.")
    else:
        print("\nFALLO: No se encontraron todos los movimientos esperados.")

if __name__ == "__main__":
    check_future_rows()
