import openpyxl
from pypdf import PdfReader
import re
import os
import sys
import datetime
from itertools import groupby

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(SCRIPT_DIR)
from config import OUTPUT_EXCEL_PATH, INPUT_PDF_DIR, log

def get_pdf_opening_balances():
    """Extract opening CP balances from PDF files for cross-verification."""
    balances = {} 
    if not os.path.exists(INPUT_PDF_DIR): return {}
    
    files = [f for f in os.listdir(INPUT_PDF_DIR) if f.lower().endswith('.pdf')]
    
    for fname in files:
        nums = re.findall(r'\d+', fname)
        if len(nums) >= 2:
            try:
                y = int(nums[-2])
                m = int(nums[-1])
                if y < 100: y += 2000
                
                fpath = os.path.join(INPUT_PDF_DIR, fname)
                reader = PdfReader(fpath)
                text = reader.pages[0].extract_text()
                
                match = re.search(r'FBA RENPEB\s+([\d\.]+,\d{2})', text)
                if match:
                    val = float(match.group(1).replace('.', '').replace(',', '.'))
                    balances[(y, m)] = val
            except Exception as e:
                log.debug(f"  No se pudo leer balance de {fname}: {e}")
            
    return balances

def verify_balances():
    log.info("=" * 50)
    log.info("PASO 4: Verificacion de Resultados")
    log.info("=" * 50)
    
    if not os.path.exists(OUTPUT_EXCEL_PATH):
        log.error(f"Archivo Excel no encontrado: {OUTPUT_EXCEL_PATH}")
        return False
    
    wb = openpyxl.load_workbook(OUTPUT_EXCEL_PATH, data_only=True)
    ws = wb.active
    
    movements = []
    for r in range(5, 10000):
        date_val = ws[f'W{r}'].value
        z_val = ws[f'Z{r}'].value
        if z_val is None: continue
        
        try:
            val = float(z_val)
            dt = None
            if isinstance(date_val, datetime.datetime):
                dt = date_val
            elif isinstance(date_val, str):
                try: dt = datetime.datetime.strptime(date_val, "%d/%m/%Y")
                except: pass
            
            if dt:
                movements.append((dt, val))
        except: pass
            
    if not movements:
        log.warning("No se encontraron movimientos en el Excel")
        return False

    # Calculate Monthly Closes
    monthly_closes = {} 
    movements.sort(key=lambda x: x[0])
    
    running_bal = 0.0
    for key, group in groupby(movements, lambda x: (x[0].year, x[0].month)):
        for m in group:
            running_bal += m[1]
        monthly_closes[key] = running_bal
        
    targets = get_pdf_opening_balances()
    
    log.info("\nReporte de Verificacion:")
    log.info("-" * 75)
    verified_count = 0
    mismatch_count = 0
    no_pdf_count = 0
    
    sorted_months = sorted(monthly_closes.keys())
    
    for y, m in sorted_months:
        ny, nm = y, m + 1
        if nm > 12:
            nm = 1
            ny += 1
            
        excel_close = monthly_closes[(y, m)]
        
        if (ny, nm) in targets:
            pdf_open = targets[(ny, nm)]
            diff = abs(excel_close - pdf_open)
            
            if diff < 0.5:
                log.info(f"  [{m:02d}/{y}] Cierre: {excel_close:>15,.2f} CPs | PDF {nm:02d}/{ny}: {pdf_open:>15,.2f} -> OK")
                verified_count += 1
            else:
                log.warning(f"  [{m:02d}/{y}] Cierre: {excel_close:>15,.2f} CPs | PDF {nm:02d}/{ny}: {pdf_open:>15,.2f} -> DIFERENCIA: {diff:.2f}")
                mismatch_count += 1
        else:
            log.info(f"  [{m:02d}/{y}] Cierre: {excel_close:>15,.2f} CPs | (Sin PDF para {nm:02d}/{ny})")
            no_pdf_count += 1
    
    log.info("-" * 75)
    log.info(f"  Coincidencias: {verified_count} | Diferencias: {mismatch_count} | Sin PDF: {no_pdf_count}")
    
    if mismatch_count > 0:
        log.warning(f"\n  ATENCION: {mismatch_count} mes(es) con diferencias de balance")
    else:
        log.info("  Todos los balances verificados correctamente")
    
    return mismatch_count == 0

if __name__ == "__main__":
    verify_balances()
