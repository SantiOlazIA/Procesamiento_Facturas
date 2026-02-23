"""
Batch transform OLD format files (2002, 2003, 2004) and merge them with the
existing 2001 transformed data into a single All_Sales_Report.xlsx.

This script:
1. Renames the existing 2001 output to All_Sales_Report.xlsx
2. Transforms each new file using the same rate-based classification logic
3. Merges all data, re-numbers, and saves
4. Reports every eliminated row with its original sheet and row position

Usage:
    python batch_transform_and_merge.py
"""
import pandas as pd
import numpy as np
import re
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# -------------------------------------------------------------------
# Configuration
# -------------------------------------------------------------------
EXISTING_REPORT = r"data\output\Libro_IVA_VTAS_2001_transformed.xlsx"
OUTPUT_FILE = r"data\output\All_Sales_Report.xlsx"

INPUT_FILES = [
    r"data\input\Libro IVA VTAS 2002.xls",
    r"data\input\Libro IVA VTAS 2003.xls",
    r"data\input\Libro IVA VTAS 2004.xls",
]

# Standard Argentine IVA rates
IVA_RATES = {0.21: "21%", 0.105: "10.5%", 0.27: "27%"}
RATE_DEVIATION_THRESHOLD = 0.01
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# New format column names
NEW_COLUMNS = [
    "Nro", "Fecha", "Tipo", "Comprobante", "Comprobante_dup", "CODIGO",
    "Razon Social", "Cuit / DNI", "Condic.",
    "A_Neto_21", "A_Neto_10_5", "B_Neto_21", "B_Neto_10_5", "IVA_Exento",
    "A_IVA_21", "A_IVA_10_5", "B_IVA_21", "B_IVA_10_5",
    "Reten_Percep", "Total",
]

COL_INDEX = {name: i + 1 for i, name in enumerate(NEW_COLUMNS)}


# -------------------------------------------------------------------
# Functions (same as transform_old_to_new.py)
# -------------------------------------------------------------------

def classify_iva_rate(neto, iva):
    if pd.isna(neto) or pd.isna(iva) or neto == 0:
        return None, None, None, False
    neto, iva = float(neto), float(iva)
    if neto == 0:
        return None, None, None, False
    effective_rate = abs(iva / neto)
    closest_rate = min(IVA_RATES.keys(), key=lambda r: abs(effective_rate - r))
    deviation = abs(effective_rate - closest_rate)
    return closest_rate, effective_rate, deviation, deviation > RATE_DEVIATION_THRESHOLD


def classify_percepcion_rate(percepcion, neto_total):
    if pd.isna(percepcion) or pd.isna(neto_total) or neto_total == 0:
        return None, False
    percepcion, neto_total = float(percepcion), float(neto_total)
    if percepcion == 0 or neto_total == 0:
        return 0.0, False
    effective_rate = abs(percepcion / neto_total)
    return effective_rate, effective_rate > 0.10


def determine_tipo(row):
    col3_val = str(row[3]).strip().upper() if pd.notna(row[3]) else ""
    return "N/C" if col3_val == "NDC" else "FAC"


def format_comprobante(row):
    tipo_letter = str(row[2]).strip().upper() if pd.notna(row[2]) else "A"
    if tipo_letter in ("NC",):
        tipo_letter = "A"
    comprobante_raw = str(row[4]).strip() if pd.notna(row[4]) else ""
    if not comprobante_raw:
        return ""
    if re.match(r'^[A-Z]\d{13}$', comprobante_raw):
        return comprobante_raw
    match = re.match(r'^(\d{4})-(\d{8})$', comprobante_raw)
    if match:
        return f"{tipo_letter}{match.group(1)}{match.group(2)}"
    match = re.match(r'^([A-Z])(\d{4})(\d{8})$', comprobante_raw)
    if match:
        return comprobante_raw
    try:
        num = int(float(comprobante_raw))
        return f"{tipo_letter}0001{num:08d}"
    except (ValueError, OverflowError):
        print(f"  WARNING: Could not parse comprobante '{comprobante_raw}'")
        return f"{tipo_letter}{comprobante_raw}"


def clean_cuit(cuit_val):
    if pd.isna(cuit_val):
        return np.nan
    return str(cuit_val).replace("-", "").strip()


def safe_numeric(val):
    if pd.isna(val):
        return np.nan
    try:
        return float(val)
    except (ValueError, TypeError):
        return np.nan


# -------------------------------------------------------------------
# Transform a single file
# -------------------------------------------------------------------

def transform_file(input_file):
    """
    Transform an old-format file into the new-format DataFrame.
    Returns: (DataFrame, list_of_highlight_cells, list_of_removed_rows)
    """
    xls = pd.ExcelFile(input_file)
    basename = os.path.basename(input_file)
    print(f"\n{'=' * 70}")
    print(f"PROCESSING: {basename}")
    print(f"Sheets: {xls.sheet_names}")
    print(f"{'=' * 70}")

    all_rows = []
    highlight_cells = []
    removed_rows = []
    skipped_sheets = []
    processed_sheets = []

    for sheet_name in xls.sheet_names:
        df = pd.read_excel(input_file, sheet_name=sheet_name, header=None)
        data = df.iloc[6:].copy()
        data = data[data[5].astype(str).str.upper() != "TOTALES"]
        data = data[~(data[1].isna() & data[9].isna())]

        if data.empty:
            skipped_sheets.append(sheet_name)
            continue

        processed_sheets.append(sheet_name)
        print(f"  Sheet '{sheet_name}': {len(data)} data rows")

        for orig_excel_idx, row in data.iterrows():
            row_idx = len(all_rows)
            # orig_excel_idx is the 0-based pandas index = Excel row (0-based)
            # Excel row for user = orig_excel_idx + 1 (1-based)
            orig_row_ref = f"Sheet '{sheet_name}', Excel row {orig_excel_idx + 1}"

            new_row = {}
            new_row["_source_file"] = basename
            new_row["_orig_ref"] = orig_row_ref

            # Basic fields
            new_row["Fecha"] = row[1]
            new_row["Tipo"] = determine_tipo(row)
            comp = format_comprobante(row)
            new_row["Comprobante"] = comp
            new_row["Comprobante_dup"] = comp
            new_row["CODIGO"] = np.nan
            new_row["Razon Social"] = row[5] if pd.notna(row[5]) else np.nan
            new_row["Cuit / DNI"] = clean_cuit(row[6])
            new_row["Condic."] = row[7] if pd.notna(row[7]) else np.nan

            # Initialize Neto/IVA
            new_row["A_Neto_21"] = np.nan
            new_row["A_Neto_10_5"] = np.nan
            new_row["B_Neto_21"] = np.nan
            new_row["B_Neto_10_5"] = np.nan
            new_row["IVA_Exento"] = safe_numeric(row[10])
            new_row["A_IVA_21"] = np.nan
            new_row["A_IVA_10_5"] = np.nan
            new_row["B_IVA_21"] = 0
            new_row["B_IVA_10_5"] = 0

            # Rate-based classification: main Neto/IVA (col 9 / col 11)
            neto_main = safe_numeric(row[9])
            iva_main = safe_numeric(row[11])

            if pd.notna(neto_main) and pd.notna(iva_main) and neto_main != 0:
                closest, eff, dev, flagged = classify_iva_rate(neto_main, iva_main)
                if closest == 0.21:
                    new_row["A_Neto_21"] = neto_main
                    new_row["A_IVA_21"] = iva_main
                    if flagged:
                        highlight_cells.append((row_idx, "A_Neto_21",
                            f"IVA rate={eff:.4f}, expected 0.21, dev={dev:.4f}"))
                        highlight_cells.append((row_idx, "A_IVA_21",
                            f"IVA rate={eff:.4f}, expected 0.21, dev={dev:.4f}"))
                elif closest == 0.105:
                    new_row["A_Neto_10_5"] = neto_main
                    new_row["A_IVA_10_5"] = iva_main
                    if flagged:
                        highlight_cells.append((row_idx, "A_Neto_10_5",
                            f"IVA rate={eff:.4f}, expected 0.105, dev={dev:.4f}"))
                        highlight_cells.append((row_idx, "A_IVA_10_5",
                            f"IVA rate={eff:.4f}, expected 0.105, dev={dev:.4f}"))
                elif closest == 0.27:
                    new_row["A_Neto_21"] = neto_main
                    new_row["A_IVA_21"] = iva_main
                    highlight_cells.append((row_idx, "A_Neto_21",
                        f"27% rate ({eff:.4f}), placed in 21% col"))
                    highlight_cells.append((row_idx, "A_IVA_21",
                        f"27% rate ({eff:.4f}), placed in 21% col"))
            elif pd.notna(neto_main):
                new_row["A_Neto_21"] = neto_main
                new_row["A_IVA_21"] = iva_main if pd.notna(iva_main) else 0

            # Rate-based classification: secondary Neto/IVA (col 8 / col 12)
            neto_sec = safe_numeric(row[8])
            iva_sec = safe_numeric(row[12])

            if pd.notna(neto_sec) and pd.notna(iva_sec) and neto_sec != 0:
                closest, eff, dev, flagged = classify_iva_rate(neto_sec, iva_sec)
                if closest == 0.105:
                    new_row["A_Neto_10_5"] = neto_sec
                    new_row["A_IVA_10_5"] = iva_sec
                    if flagged:
                        highlight_cells.append((row_idx, "A_Neto_10_5",
                            f"Sec. rate={eff:.4f}, expected 0.105, dev={dev:.4f}"))
                        highlight_cells.append((row_idx, "A_IVA_10_5",
                            f"Sec. rate={eff:.4f}, expected 0.105, dev={dev:.4f}"))
                elif closest == 0.21:
                    if pd.isna(new_row["A_Neto_21"]):
                        new_row["A_Neto_21"] = neto_sec
                        new_row["A_IVA_21"] = iva_sec
                    else:
                        new_row["A_Neto_21"] = float(new_row["A_Neto_21"]) + float(neto_sec)
                        new_row["A_IVA_21"] = float(new_row["A_IVA_21"]) + float(iva_sec)
                    highlight_cells.append((row_idx, "A_Neto_21",
                        f"Secondary neto classified as 21% (rate={eff:.4f})"))
                else:
                    new_row["A_Neto_10_5"] = neto_sec
                    new_row["A_IVA_10_5"] = iva_sec
            elif pd.notna(neto_sec):
                new_row["A_Neto_10_5"] = neto_sec
                new_row["A_IVA_10_5"] = iva_sec if pd.notna(iva_sec) else 0

            # Percepciones
            percep = safe_numeric(row[13])
            new_row["Reten_Percep"] = percep
            total_neto = sum(abs(float(n)) for n in [neto_main, neto_sec] if pd.notna(n))
            if pd.notna(percep) and percep != 0 and total_neto > 0:
                percep_rate, percep_flagged = classify_percepcion_rate(percep, total_neto)
                if percep_flagged:
                    highlight_cells.append((row_idx, "Reten_Percep",
                        f"Percep. rate={percep_rate:.4f} ({percep_rate*100:.1f}% of neto)"))

            # Total
            new_row["Total"] = safe_numeric(row[14])

            # Check if Total is 0 -> mark for removal
            total_val = new_row["Total"]
            if pd.isna(total_val) or total_val == 0:
                removed_rows.append({
                    "source_file": basename,
                    "orig_ref": orig_row_ref,
                    "comprobante": comp,
                    "razon_social": new_row["Razon Social"],
                    "total": total_val,
                })
            else:
                all_rows.append(new_row)

    print(f"\n  Summary for {basename}:")
    print(f"    Skipped sheets: {skipped_sheets}")
    print(f"    Processed sheets: {processed_sheets}")
    print(f"    Data rows kept: {len(all_rows)}")
    print(f"    Rows removed (Total=0): {len(removed_rows)}")
    print(f"    Cells flagged: {len(highlight_cells)}")

    if all_rows:
        result_df = pd.DataFrame(all_rows)
    else:
        result_df = pd.DataFrame(columns=NEW_COLUMNS + ["_source_file", "_orig_ref"])

    return result_df, highlight_cells, removed_rows


# -------------------------------------------------------------------
# Main
# -------------------------------------------------------------------

def main():
    # Step 1: Load existing 2001 data
    if os.path.exists(EXISTING_REPORT):
        print(f"Loading existing data from: {EXISTING_REPORT}")
        existing_df = pd.read_excel(EXISTING_REPORT)
        existing_df["_source_file"] = "Libro IVA VTAS 2001.xls"
        existing_df["_orig_ref"] = "Previously transformed"
        print(f"  Existing rows: {len(existing_df)}")
    else:
        print(f"WARNING: {EXISTING_REPORT} not found, starting fresh.")
        existing_df = pd.DataFrame(columns=NEW_COLUMNS + ["_source_file", "_orig_ref"])

    # Step 2: Transform each new file
    all_new_dfs = []
    all_highlights = []
    all_removed = []
    highlight_offset = len(existing_df)  # offset for highlight row indices

    for input_file in INPUT_FILES:
        if not os.path.exists(input_file):
            print(f"ERROR: File not found: {input_file}")
            sys.exit(1)

        file_df, highlights, removed = transform_file(input_file)
        
        # Offset highlight indices by the accumulated row count
        adjusted_highlights = [
            (idx + highlight_offset, col, reason)
            for idx, col, reason in highlights
        ]
        all_highlights.extend(adjusted_highlights)
        all_removed.extend(removed)
        
        all_new_dfs.append(file_df)
        highlight_offset += len(file_df)

    # Step 3: Merge all DataFrames
    frames = [existing_df] + all_new_dfs
    merged_df = pd.concat(frames, ignore_index=True)

    # Ensure column order and drop internal tracking columns
    for col in NEW_COLUMNS:
        if col not in merged_df.columns:
            merged_df[col] = np.nan

    # Re-number
    merged_df["Nro"] = range(1, len(merged_df) + 1)

    # Keep only the output columns (drop _source_file, _orig_ref)
    output_df = merged_df[NEW_COLUMNS].copy()

    # Step 4: Print final summary
    print(f"\n{'=' * 70}")
    print(f"MERGE COMPLETE")
    print(f"{'=' * 70}")
    print(f"  Total rows in merged report: {len(output_df)}")
    print(f"  Date range: {merged_df['Fecha'].min()} to {merged_df['Fecha'].max()}")
    print(f"  Tipo breakdown: {output_df['Tipo'].value_counts().to_dict()}")
    print(f"  Total yellow flags: {len(all_highlights)}")

    # Print ALL removed rows with original references
    if all_removed:
        print(f"\n--- REMOVED ROWS (Total=0) ---")
        print(f"  Total removed: {len(all_removed)}")
        for i, r in enumerate(all_removed, 1):
            print(f"  {i}. File: {r['source_file']}")
            print(f"     Original position: {r['orig_ref']}")
            print(f"     Comprobante: {r['comprobante']}")
            print(f"     Razon Social: {r['razon_social']}")
            print(f"     Total: {r['total']}")
    else:
        print(f"\n  No rows removed (Total=0) from the new files.")

    # Print flagged cells
    if all_highlights:
        print(f"\n--- FLAGGED CELLS (yellow highlight) ---")
        for row_idx, col_name, reason in all_highlights:
            print(f"  Row {row_idx+1}, col '{col_name}': {reason}")

    # Step 5: Save to Excel
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    output_df.to_excel(OUTPUT_FILE, index=False, sheet_name="ventas")
    print(f"\n  Saved to: {OUTPUT_FILE}")

    # Step 6: Apply highlights
    if all_highlights:
        wb = load_workbook(OUTPUT_FILE)
        ws = wb["ventas"]
        applied = 0
        for row_idx, col_name, reason in all_highlights:
            excel_row = row_idx + 2  # +1 for 1-based, +1 for header
            excel_col = COL_INDEX[col_name]
            ws.cell(row=excel_row, column=excel_col).fill = YELLOW_FILL
            applied += 1
        wb.save(OUTPUT_FILE)
        print(f"  Applied {applied} yellow highlights.")

    # Show first and last few rows
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', 250)
    pd.set_option('display.max_colwidth', 35)
    print(f"\n--- First 3 rows ---")
    print(output_df.head(3).to_string(index=False))
    print(f"\n--- Last 3 rows ---")
    print(output_df.tail(3).to_string(index=False))


if __name__ == "__main__":
    main()
