"""
Transform OLD format (Libro IVA VTAS 2001.xls) to NEW format (20-column layout).
This script is READ-ONLY on the input file. Output goes to data/output/.

Features:
  - Rate-based IVA classification: computes IVA/Neto, assigns to closest rate column
  - Yellow highlighting for significant deviations from expected rates
  - Same analysis for percepciones/retenciones
  - Removes comprobantes with Total = 0

Usage:
    python transform_old_to_new.py
"""
import pandas as pd
import numpy as np
import re
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# -------------------------------------------------------------------
# Configuration
# -------------------------------------------------------------------
INPUT_FILE = r"data\input\Libro IVA VTAS 2001.xls"
OUTPUT_FILE = r"data\output\Libro_IVA_VTAS_2001_transformed.xlsx"

# Standard Argentine IVA rates
IVA_RATES = {
    0.21:  "21%",
    0.105: "10.5%",
    0.27:  "27%",
}

# Deviation threshold: flag if effective rate differs from closest standard
# rate by more than this (as absolute difference, e.g. 0.01 = 1 percentage point)
RATE_DEVIATION_THRESHOLD = 0.01

# Yellow fill for flagged cells
YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

# New format column names (matching row 5 of the new-format file)
NEW_COLUMNS = [
    "Nro",                  # 0  - Sequential row number
    "Fecha",                # 1  - Date
    "Tipo",                 # 2  - FAC / N/C
    "Comprobante",          # 3  - Formatted comprobante
    "Comprobante_dup",      # 4  - Duplicate of comprobante
    "CODIGO",               # 5  - Client code (not in old format)
    "Razon Social",         # 6  - Client name
    "Cuit / DNI",           # 7  - CUIT without hyphens
    "Condic.",              # 8  - IVA condition
    "A_Neto_21",            # 9  - "A" Neto 21%
    "A_Neto_10_5",          # 10 - "A" Neto 10,5%
    "B_Neto_21",            # 11 - "B" Neto 21%
    "B_Neto_10_5",          # 12 - "B" Neto 10,5%
    "IVA_Exento",           # 13 - IVA Exento
    "A_IVA_21",             # 14 - "A" IVA 21%
    "A_IVA_10_5",           # 15 - "A" IVA 10,5%
    "B_IVA_21",             # 16 - "B" IVA 21%
    "B_IVA_10_5",           # 17 - "B" IVA 10,5%
    "Reten_Percep",         # 18 - Retenciones / Percepciones
    "Total",                # 19 - Total
]

# Column name -> 1-based Excel column index mapping (for highlighting)
COL_INDEX = {name: i + 1 for i, name in enumerate(NEW_COLUMNS)}


# -------------------------------------------------------------------
# Rate Classification Functions
# -------------------------------------------------------------------

def classify_iva_rate(neto, iva):
    """
    Given a Neto and IVA amount, compute the effective rate and determine
    which standard rate (21%, 10.5%, 27%) is the closest match.

    Returns:
        (closest_rate, effective_rate, deviation, is_flagged)
        - closest_rate: float (0.21, 0.105, or 0.27)
        - effective_rate: float (actual IVA/Neto)
        - deviation: absolute difference from closest rate
        - is_flagged: True if deviation > threshold
    """
    if pd.isna(neto) or pd.isna(iva) or neto == 0:
        return None, None, None, False

    neto = float(neto)
    iva = float(iva)

    if neto == 0:
        return None, None, None, False

    effective_rate = abs(iva / neto)

    # Find closest standard rate
    closest_rate = min(IVA_RATES.keys(), key=lambda r: abs(effective_rate - r))
    deviation = abs(effective_rate - closest_rate)
    is_flagged = deviation > RATE_DEVIATION_THRESHOLD

    return closest_rate, effective_rate, deviation, is_flagged


def classify_percepcion_rate(percepcion, neto_total):
    """
    Analyze percepciones/retenciones relative to the total neto.
    Common perception rates in Argentina: 1%, 1.5%, 3%, 5%, 6%

    Returns:
        (effective_rate, is_flagged)
        - effective_rate: percepcion / neto_total
        - is_flagged: True if rate seems unusual (>10% or negative)
    """
    if pd.isna(percepcion) or pd.isna(neto_total) or neto_total == 0:
        return None, False

    percepcion = float(percepcion)
    neto_total = float(neto_total)

    if percepcion == 0 or neto_total == 0:
        return 0.0, False

    effective_rate = abs(percepcion / neto_total)

    # Flag if the perception rate is > 10% of neto (unusually high)
    # or if the sign is unexpected
    is_flagged = effective_rate > 0.10

    return effective_rate, is_flagged


# -------------------------------------------------------------------
# Transformation Functions
# -------------------------------------------------------------------

def determine_tipo(row):
    """
    Determine the Tipo field for the new format.
    Old format: col2 = letter (A), col3 = 1 (FAC) or NDC (Nota de Credito)
    New format: FAC or N/C
    """
    col3_val = str(row[3]).strip().upper() if pd.notna(row[3]) else ""
    if col3_val == "NDC":
        return "N/C"
    else:
        return "FAC"


def format_comprobante(row):
    """
    Format the comprobante number from old to new format.
    Old: "0001-00000100" -> New: "A000100000100"
    Old (bare number): "635" -> "A000100000635" (padded)
    """
    tipo_letter = str(row[2]).strip().upper() if pd.notna(row[2]) else "A"
    if tipo_letter in ("NC",):
        tipo_letter = "A"

    comprobante_raw = str(row[4]).strip() if pd.notna(row[4]) else ""

    if not comprobante_raw:
        return ""

    # Already formatted like "A000100000100"
    if re.match(r'^[A-Z]\d{13}$', comprobante_raw):
        return comprobante_raw

    # Format: "0001-00000100"
    match = re.match(r'^(\d{4})-(\d{8})$', comprobante_raw)
    if match:
        return f"{tipo_letter}{match.group(1)}{match.group(2)}"

    # Format: "A000100000100" already with prefix
    match = re.match(r'^([A-Z])(\d{4})(\d{8})$', comprobante_raw)
    if match:
        return comprobante_raw

    # Bare number (e.g., "635", "1265")
    try:
        num = int(float(comprobante_raw))
        return f"{tipo_letter}0001{num:08d}"
    except (ValueError, OverflowError):
        print(f"  WARNING: Could not parse comprobante '{comprobante_raw}', keeping as-is")
        return f"{tipo_letter}{comprobante_raw}"


def clean_cuit(cuit_val):
    """Remove hyphens from CUIT. Old: '30-70756854-9' -> '30707568549'"""
    if pd.isna(cuit_val):
        return np.nan
    return str(cuit_val).replace("-", "").strip()


def safe_numeric(val):
    """Safely convert a value to numeric, returning NaN if not possible."""
    if pd.isna(val):
        return np.nan
    try:
        return float(val)
    except (ValueError, TypeError):
        return np.nan


def remove_zero_total_comprobantes(df):
    """
    Remove rows where the Total is exactly 0.
    These are typically voided/annulled comprobantes (COMPROBANTE ANULADO).
    """
    mask = df["Total"].fillna(0) == 0
    removed = df[mask]
    result = df[~mask]
    removed_count = len(removed)

    if removed_count > 0:
        print(f"\n  Removed {removed_count} comprobante(s) with Total = 0:")
        for _, r in removed.iterrows():
            comp = r['Comprobante']
            name = r['Razon Social']
            print(f"    - Nro {r['Nro']}: {comp} | {name} | Total: {r['Total']}")

    return result, removed_count


# -------------------------------------------------------------------
# Main Processing
# -------------------------------------------------------------------

def process_file():
    """Main function to transform old format to new format."""

    if not os.path.exists(INPUT_FILE):
        print(f"ERROR: Input file not found: {INPUT_FILE}")
        sys.exit(1)

    xls = pd.ExcelFile(INPUT_FILE)
    print(f"Input file: {INPUT_FILE}")
    print(f"Sheets found: {xls.sheet_names}")

    all_rows = []
    # Track cells to highlight: list of (row_index_in_all_rows, column_name, reason)
    highlight_cells = []
    skipped_sheets = []
    processed_sheets = []

    for sheet_name in xls.sheet_names:
        df = pd.read_excel(INPUT_FILE, sheet_name=sheet_name, header=None)

        # Skip header rows (0-5) - data starts at row 6
        data = df.iloc[6:].copy()

        # Filter out TOTALES
        data = data[data[5].astype(str).str.upper() != "TOTALES"]

        # Filter out empty filler rows
        data = data[~(data[1].isna() & data[9].isna())]

        if data.empty:
            skipped_sheets.append(sheet_name)
            continue

        processed_sheets.append(sheet_name)
        print(f"\n  Processing sheet '{sheet_name}': {len(data)} data rows")

        for _, row in data.iterrows():
            row_idx = len(all_rows)  # 0-based index in the output
            new_row = {}

            # --- Basic fields ---
            new_row["Fecha"] = row[1]
            new_row["Tipo"] = determine_tipo(row)
            comp = format_comprobante(row)
            new_row["Comprobante"] = comp
            new_row["Comprobante_dup"] = comp
            new_row["CODIGO"] = np.nan
            new_row["Razon Social"] = row[5] if pd.notna(row[5]) else np.nan
            new_row["Cuit / DNI"] = clean_cuit(row[6])
            new_row["Condic."] = row[7] if pd.notna(row[7]) else np.nan

            # --- Initialize all Neto/IVA columns ---
            new_row["A_Neto_21"] = np.nan
            new_row["A_Neto_10_5"] = np.nan
            new_row["B_Neto_21"] = np.nan
            new_row["B_Neto_10_5"] = np.nan
            new_row["IVA_Exento"] = safe_numeric(row[10])
            new_row["A_IVA_21"] = np.nan
            new_row["A_IVA_10_5"] = np.nan
            new_row["B_IVA_21"] = 0
            new_row["B_IVA_10_5"] = 0

            # --- Rate-based IVA classification for main Neto/IVA pair ---
            # Old col 9 = Neto (main), Old col 11 = IVA General
            neto_main = safe_numeric(row[9])
            iva_main = safe_numeric(row[11])

            if pd.notna(neto_main) and pd.notna(iva_main) and neto_main != 0:
                closest_rate, eff_rate, deviation, flagged = classify_iva_rate(
                    neto_main, iva_main
                )

                if closest_rate == 0.21:
                    new_row["A_Neto_21"] = neto_main
                    new_row["A_IVA_21"] = iva_main
                    if flagged:
                        highlight_cells.append(
                            (row_idx, "A_Neto_21",
                             f"IVA rate={eff_rate:.4f}, expected 0.21, dev={deviation:.4f}")
                        )
                        highlight_cells.append(
                            (row_idx, "A_IVA_21",
                             f"IVA rate={eff_rate:.4f}, expected 0.21, dev={deviation:.4f}")
                        )
                elif closest_rate == 0.105:
                    new_row["A_Neto_10_5"] = neto_main
                    new_row["A_IVA_10_5"] = iva_main
                    if flagged:
                        highlight_cells.append(
                            (row_idx, "A_Neto_10_5",
                             f"IVA rate={eff_rate:.4f}, expected 0.105, dev={deviation:.4f}")
                        )
                        highlight_cells.append(
                            (row_idx, "A_IVA_10_5",
                             f"IVA rate={eff_rate:.4f}, expected 0.105, dev={deviation:.4f}")
                        )
                elif closest_rate == 0.27:
                    # 27% goes into the 21% column (closest available) but flagged
                    new_row["A_Neto_21"] = neto_main
                    new_row["A_IVA_21"] = iva_main
                    highlight_cells.append(
                        (row_idx, "A_Neto_21",
                         f"Detected 27% rate ({eff_rate:.4f}), placed in 21% col")
                    )
                    highlight_cells.append(
                        (row_idx, "A_IVA_21",
                         f"Detected 27% rate ({eff_rate:.4f}), placed in 21% col")
                    )
            elif pd.notna(neto_main):
                # Has Neto but no IVA — could be exempt, assign to 21% by default
                new_row["A_Neto_21"] = neto_main
                new_row["A_IVA_21"] = iva_main if pd.notna(iva_main) else 0

            # --- Rate-based classification for secondary Neto/IVA pair ---
            # Old col 8 = Neto (B y Z), Old col 12 = IVA No Inscripto
            neto_sec = safe_numeric(row[8])
            iva_sec = safe_numeric(row[12])

            if pd.notna(neto_sec) and pd.notna(iva_sec) and neto_sec != 0:
                closest_rate, eff_rate, deviation, flagged = classify_iva_rate(
                    neto_sec, iva_sec
                )

                if closest_rate == 0.105:
                    new_row["A_Neto_10_5"] = neto_sec
                    new_row["A_IVA_10_5"] = iva_sec
                    if flagged:
                        highlight_cells.append(
                            (row_idx, "A_Neto_10_5",
                             f"Sec. IVA rate={eff_rate:.4f}, expected 0.105, dev={deviation:.4f}")
                        )
                        highlight_cells.append(
                            (row_idx, "A_IVA_10_5",
                             f"Sec. IVA rate={eff_rate:.4f}, expected 0.105, dev={deviation:.4f}")
                        )
                elif closest_rate == 0.21:
                    # Secondary at 21%? Unusual — add to 21% but flag
                    if pd.isna(new_row["A_Neto_21"]):
                        new_row["A_Neto_21"] = neto_sec
                        new_row["A_IVA_21"] = iva_sec
                    else:
                        # Already have 21% values — add secondary amounts
                        new_row["A_Neto_21"] = float(new_row["A_Neto_21"]) + float(neto_sec)
                        new_row["A_IVA_21"] = float(new_row["A_IVA_21"]) + float(iva_sec)
                    highlight_cells.append(
                        (row_idx, "A_Neto_21",
                         f"Secondary neto classified as 21% (rate={eff_rate:.4f})")
                    )
                else:
                    # Default: put in 10.5%
                    new_row["A_Neto_10_5"] = neto_sec
                    new_row["A_IVA_10_5"] = iva_sec
            elif pd.notna(neto_sec):
                # Has secondary Neto but no IVA
                new_row["A_Neto_10_5"] = neto_sec
                new_row["A_IVA_10_5"] = iva_sec if pd.notna(iva_sec) else 0

            # --- Percepciones / Retenciones ---
            percep = safe_numeric(row[13])
            new_row["Reten_Percep"] = percep

            # Analyze perception rate relative to total neto
            total_neto = 0
            for n in [neto_main, neto_sec]:
                if pd.notna(n):
                    total_neto += abs(float(n))

            if pd.notna(percep) and percep != 0 and total_neto > 0:
                percep_rate, percep_flagged = classify_percepcion_rate(
                    percep, total_neto
                )
                if percep_flagged:
                    highlight_cells.append(
                        (row_idx, "Reten_Percep",
                         f"Perception rate={percep_rate:.4f} ({percep_rate*100:.1f}% of neto), unusually high")
                    )

            # --- Total ---
            new_row["Total"] = safe_numeric(row[14])

            all_rows.append(new_row)

    # Build final DataFrame
    if not all_rows:
        print("ERROR: No data rows found across any sheets!")
        sys.exit(1)

    result_df = pd.DataFrame(all_rows)
    result_df.insert(0, "Nro", range(1, len(result_df) + 1))
    result_df = result_df[NEW_COLUMNS]

    # Remove comprobantes with Total = 0
    result_df, zero_count = remove_zero_total_comprobantes(result_df)
    result_df["Nro"] = range(1, len(result_df) + 1)

    # --- Print summary ---
    print(f"\n{'=' * 60}")
    print(f"TRANSFORMATION COMPLETE")
    print(f"{'=' * 60}")
    print(f"  Skipped sheets (no data): {skipped_sheets}")
    print(f"  Processed sheets: {processed_sheets}")
    print(f"  Total rows in output: {len(result_df)}")
    print(f"  Rows removed (Total=0): {zero_count}")
    print(f"  Tipo breakdown: {result_df['Tipo'].value_counts().to_dict()}")
    print(f"  Date range: {result_df['Fecha'].min()} to {result_df['Fecha'].max()}")
    print(f"  Cells flagged (yellow): {len(highlight_cells)}")

    # Show first 5 rows
    pd.set_option('display.max_columns', None)
    pd.set_option('display.width', 250)
    pd.set_option('display.max_colwidth', 35)
    print(f"\n--- First 5 rows of output ---")
    print(result_df.head().to_string(index=False))

    # Show flagged items
    if highlight_cells:
        print(f"\n--- Flagged cells (will be highlighted yellow) ---")
        for row_idx, col_name, reason in highlight_cells:
            comp = all_rows[row_idx].get("Comprobante", "?")
            print(f"  Row {row_idx+1} [{comp}] col '{col_name}': {reason}")

    # --- Save to Excel ---
    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    result_df.to_excel(OUTPUT_FILE, index=False, sheet_name="ventas transformadas")

    # --- Apply yellow highlighting with openpyxl ---
    if highlight_cells:
        wb = load_workbook(OUTPUT_FILE)
        ws = wb["ventas transformadas"]

        # Build a set of removed row indices for offset calculation
        # After remove_zero_total, the DataFrame was re-indexed, so we need
        # to map original row_idx to final row position
        # The highlight_cells use the original all_rows index, so we need
        # to check which rows survived the zero-total filter
        zero_mask = pd.DataFrame(all_rows)["Total"].fillna(0) == 0
        surviving_indices = [i for i in range(len(all_rows)) if not zero_mask.iloc[i]]
        idx_to_excel_row = {orig_idx: excel_row for excel_row, orig_idx
                           in enumerate(surviving_indices, start=2)}  # +2 for 1-based + header

        applied_count = 0
        for row_idx, col_name, reason in highlight_cells:
            if row_idx in idx_to_excel_row:
                excel_row = idx_to_excel_row[row_idx]
                excel_col = COL_INDEX[col_name]
                ws.cell(row=excel_row, column=excel_col).fill = YELLOW_FILL
                applied_count += 1

        wb.save(OUTPUT_FILE)
        print(f"\n  Applied {applied_count} yellow highlights to Excel.")

    print(f"\n  Output saved to: {OUTPUT_FILE}")


if __name__ == "__main__":
    process_file()
