import pandas as pd
import numpy as np
import os
from openpyxl import Workbook
from openpyxl.chart import LineChart, Reference, BarChart, AreaChart
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font, Color, PatternFill, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule

# -------------------------------------------------------------------
# Configuration
# -------------------------------------------------------------------
SOURCE_FILE = r'data\output\All_Sales_Report.xlsx'
OUTPUT_ANALYTICAL = r'data\output\Analytical_Sales_Reports.xlsx'

if not os.path.exists(SOURCE_FILE):
    print(f"Error: {SOURCE_FILE} not found.")
    exit(1)

# -------------------------------------------------------------------
# Styles (Premium Palette)
# -------------------------------------------------------------------
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid") # Deep Blue
HEADER_FONT = Font(color="FFFFFF", bold=True)
THIN_SIDE = Side(style='thin', color="BFBFBF")
THIN_BORDER = Border(left=THIN_SIDE, right=THIN_SIDE, top=THIN_SIDE, bottom=THIN_SIDE)
CENTER_ALIGN = Alignment(horizontal="center", vertical="center")
NUM_FORMAT = '#,##0.00'
PCT_FORMAT = '0.0%'

def apply_table_style(ws, min_row, max_row, min_col, max_col):
    """Apply borders and centering to a range of cells."""
    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            cell.border = THIN_BORDER
            cell.alignment = CENTER_ALIGN

from openpyxl.utils import get_column_letter

def auto_fit_columns(ws, max_width=40):
    """Adjust column widths based on content, with a cap."""
    for col_idx, col in enumerate(ws.columns, 1):
        max_length = 0
        column_letter = get_column_letter(col_idx)
        for cell in col:
            # Skip merged cells for length calculation, use their raw value if possible
            if cell.value:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
        adjusted_width = min(max_length + 2, max_width)
        ws.column_dimensions[column_letter].width = adjusted_width

# -------------------------------------------------------------------
# Data Preparation
# -------------------------------------------------------------------
print("Loading data for premium restructuring...")
df = pd.read_excel(SOURCE_FILE)
df['Fecha'] = pd.to_datetime(df['Fecha'])
df['Year'] = df['Fecha'].dt.year.fillna(0).astype(int)

# Calculate Total Neto (handling possible NaNs)
neto_cols = [c for c in df.columns if 'Neto' in c]
df['Neto_Total'] = df[neto_cols].sum(axis=1)

# Identify top 15 by Total cumulative
top_stats = df.groupby('Razon Social')['Total'].sum().nlargest(15).reset_index()
top_clients = top_stats['Razon Social'].tolist()
top_stats['Label'] = top_stats['Razon Social'].apply(lambda x: (str(x)[:20] + '..') if len(str(x)) > 22 else str(x))

# Total Company Revenue for Summary
total_company_total = df['Total'].sum()
total_company_net = df['Neto_Total'].sum()

# Aggregations for Concentration
yearly_total_company = df.groupby('Year')['Total'].sum().reset_index().rename(columns={'Total': 'Company_Total'})
yearly_top15_sum = df[df['Razon Social'].isin(top_clients)].groupby('Year')['Total'].sum().reset_index().rename(columns={'Total': 'Top15_Total'})

# Merge for Concentration Analysis
concentration_df = pd.merge(yearly_total_company, yearly_top15_sum, on='Year', how='left').fillna(0)
concentration_df['Top15_Share'] = (concentration_df['Top15_Total'] / concentration_df['Company_Total']).fillna(0)
concentration_df['Others_Share'] = 1 - concentration_df['Top15_Share']

# Yearly per-client pivots for charts
pivot_total = df[df['Razon Social'].isin(top_clients)].pivot_table(index='Year', columns='Razon Social', values='Total', aggfunc='sum').fillna(0)
pivot_neto = df[df['Razon Social'].isin(top_clients)].pivot_table(index='Year', columns='Razon Social', values='Neto_Total', aggfunc='sum').fillna(0)

# -------------------------------------------------------------------
# Workbook Orchestration
# -------------------------------------------------------------------
wb = Workbook()
wb.remove(wb.active)

ws_dash = wb.create_sheet("Home Dashboard")
ws_data = wb.create_sheet("Analytical Data")

# Professional View
ws_dash.sheet_view.showGridLines = False

# -------------------------------------------------------------------
# Sheet 1: Home Dashboard & Navigation
# -------------------------------------------------------------------
print("Creating Dashboard and Navigation...")
ws_dash.merge_cells("A1:I1")
ws_dash["A1"] = "CLIENT PORTFOLIO MANAGEMENT SYSTEM"
ws_dash["A1"].font = Font(bold=True, size=20, color="1F4E78")
ws_dash["A1"].alignment = CENTER_ALIGN

# Executive Summary KPIs
ws_dash["A3"] = "Executive Summary"
ws_dash["A3"].font = Font(bold=True, size=14)
ws_dash["A4"] = "Total Revenue (Gross):"
ws_dash["B4"] = total_company_total
ws_dash["B4"].number_format = NUM_FORMAT
ws_dash["A5"] = "Net Revenue (Excl. IVA):"
ws_dash["B5"] = total_company_net
ws_dash["B5"].number_format = NUM_FORMAT
ws_dash["A6"] = "Top 15 Clients Weight:"
ws_dash["B6"] = (df[df['Razon Social'].isin(top_clients)]['Total'].sum() / total_company_total)
ws_dash["B6"].number_format = PCT_FORMAT

apply_table_style(ws_dash, 4, 6, 1, 2)

# Navigation Menu
ws_dash["D3"] = "Jump to Individual Portfolios"
ws_dash["D3"].font = Font(bold=True, size=14)
ws_dash["D3"].alignment = CENTER_ALIGN

menu_start_row = 4
for i, client in enumerate(top_clients):
    short_name = (str(client)[:25] + "..") if len(str(client)) > 27 else str(client)
    sheet_name = f"Client_{i+1}" 
    safe_title = f"{i+1}. {short_name}"
    
    cell_pos = f"D{menu_start_row + i}"
    ws_dash[cell_pos] = safe_title
    ws_dash[cell_pos].hyperlink = f"#'{sheet_name}'!A1"
    ws_dash[cell_pos].font = Font(color="0563C1", underline="single")
    ws_dash[cell_pos].alignment = CENTER_ALIGN
    ws_dash[cell_pos].border = THIN_BORDER

# Concentration Analysis Chart (Center-Right)
chart_conc = AreaChart()
chart_conc.title = "Global Market Concentration (Top 15 vs Others)"
chart_conc.style = 48
chart_conc.y_axis.title = "Share %"
chart_conc.grouping = "stacked"
chart_conc.height = 10
chart_conc.width = 18

data_conc = Reference(ws_data, min_col=4, max_col=5, min_row=2, max_row=concentration_df.shape[0]+2)
cats_conc = Reference(ws_data, min_col=1, min_row=3, max_row=concentration_df.shape[0]+2)
chart_conc.add_data(data_conc, titles_from_data=True)
chart_conc.set_categories(cats_conc)
ws_dash.add_chart(chart_conc, "G3") # Moved slightly right to avoid menu

auto_fit_columns(ws_dash)

# -------------------------------------------------------------------
# Sheet 2: Analytical Data (Global Data Asset)
# -------------------------------------------------------------------
print("Populating and Formatting Global Data Asset...")
ws_data.append(["CONCENTRATION ANALYSIS HISTORY"])
ws_data.append(["Year", "Company Total", "Top 15 Total", "Top 15 Share %", "Others Share %"])
for r in dataframe_to_rows(concentration_df, index=False, header=False):
    ws_data.append(r)

ws_data.append([])
yearly_total_start = ws_data.max_row + 1
ws_data.append(["GLOBAL YEARLY TOTAL SALES PER CLIENT"])
for r in dataframe_to_rows(pivot_total.reset_index(), index=False, header=True):
    ws_data.append(r)

ws_data.append([])
yearly_net_start = ws_data.max_row + 1
ws_data.append(["GLOBAL YEARLY NET SALES PER CLIENT"])
for r in dataframe_to_rows(pivot_neto.reset_index(), index=False, header=True):
    ws_data.append(r)

# Premium Formatting for Sheet 2
apply_table_style(ws_data, 1, ws_data.max_row, 1, ws_data.max_column)

# Corporate Headers for all sub-tables in Data Sheet
data_header_rows = [2, yearly_total_start + 1, yearly_net_start + 1]
data_title_rows = [1, yearly_total_start, yearly_net_start]

for row_idx in data_header_rows:
    for cell in ws_data[row_idx]:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT

for row_idx in data_title_rows:
    ws_data.cell(row=row_idx, column=1).font = Font(bold=True, size=12, color="1F4E78")

auto_fit_columns(ws_data)

# -------------------------------------------------------------------
# Individual Client Sheets
# -------------------------------------------------------------------
print("Generating 15 Individual Portfolio Sheets...")
for i, client in enumerate(top_clients):
    sheet_id = f"Client_{i+1}"
    full_name = str(client)
    ws_p = wb.create_sheet(sheet_id)
    ws_p.sheet_view.showGridLines = False
    
    # Portfolio Header
    ws_p.merge_cells("A1:K1")
    ws_p["A1"] = f"PORTFOLIO ANALYSIS: {full_name}"
    ws_p["A1"].font = Font(bold=True, size=18, color="1F4E78")
    ws_p["A1"].alignment = CENTER_ALIGN
    
    # Back to Dashboard Link
    ws_p["A2"] = "← Back to Home Dashboard"
    ws_p["A2"].hyperlink = "#'Home Dashboard'!A1"
    ws_p["A2"].font = Font(color="0563C1", italic=True)
    ws_p["A2"].alignment = Alignment(horizontal="left")
    
    # Client KPIs
    client_total = df[df['Razon Social'] == client]['Total'].sum()
    client_net = df[df['Razon Social'] == client]['Neto_Total'].sum()
    
    ws_p["A4"] = "Lifetime Total Sales:"
    ws_p["B4"] = client_total
    ws_p["B4"].number_format = NUM_FORMAT
    ws_p["A5"] = "Lifetime Net Sales:"
    ws_p["B5"] = client_net
    ws_p["B5"].number_format = NUM_FORMAT
    
    # KPI Center & Border
    apply_table_style(ws_p, 4, 5, 1, 2)
    for row in ws_p["A4:B5"]:
        for cell in row: cell.font = Font(bold=True)
    
    # High-Resolution Dual Chart (Positioned Below KPIs)
    chart = LineChart()
    chart.title = f"Historical Performance (Net vs Gross)"
    chart.style = 13
    chart.height = 13
    chart.width = 25
    chart.y_axis.title = "Revenue ($)"
    chart.y_axis.majorUnit = 1000 
    chart.y_axis.scaling.min = 0 
    
    # Series 1: Total Sales
    ref_total = Reference(ws_data, min_col=i+2, min_row=yearly_total_start+1, max_row=yearly_total_start + len(pivot_total) + 1)
    chart.add_data(ref_total, titles_from_data=True)
    # Series 2: Net Sales
    ref_net = Reference(ws_data, min_col=i+2, min_row=yearly_net_start+1, max_row=yearly_net_start + len(pivot_neto) + 1)
    chart.add_data(ref_net, titles_from_data=True)
    
    cats = Reference(ws_data, min_col=1, min_row=yearly_total_start+2, max_row=yearly_total_start + len(pivot_total) + 1)
    chart.set_categories(cats)
    
    ws_p.add_chart(chart, "A8") # Down more to avoid overlaps
    
    # Trailing Table (Last years) - Positioned Below Chart
    data_start_row = 35 
    ws_p.cell(row=data_start_row-1, column=1).value = "Historical Data Detail"
    ws_p.cell(row=data_start_row-1, column=1).font = Font(bold=True, size=12)
    ws_p.cell(row=data_start_row-1, column=1).alignment = CENTER_ALIGN
    
    # Headers
    ws_p.cell(row=data_start_row, column=1).value = "Year"
    ws_p.cell(row=data_start_row, column=2).value = "Total Sales"
    ws_p.cell(row=data_start_row, column=3).value = "Net Sales"
    
    # Data Data
    for idx, year in enumerate(pivot_total.index):
        row_idx = data_start_row + 1 + idx
        ws_p.cell(row=row_idx, column=1).value = year
        ws_p.cell(row=row_idx, column=2).value = pivot_total.loc[year, client]
        ws_p.cell(row=row_idx, column=3).value = pivot_neto.loc[year, client]
        
    # Style Table
    apply_table_style(ws_p, data_start_row, data_start_row + len(pivot_total), 1, 3)
    for row in ws_p.iter_rows(min_row=data_start_row + 1, max_row=data_start_row + len(pivot_total), min_col=2, max_col=3):
        for cell in row: cell.number_format = NUM_FORMAT
    # Header style
    for col_idx in range(1, 4):
        ws_p.cell(row=data_start_row, column=col_idx).fill = HEADER_FILL
        ws_p.cell(row=data_start_row, column=col_idx).font = HEADER_FONT

    auto_fit_columns(ws_p)

# -------------------------------------------------------------------
# Final Polish
# -------------------------------------------------------------------
# Formatting technical data sheet
for row in ws_data.iter_rows(min_row=3, max_row=ws_data.max_row, min_col=2, max_col=ws_data.max_column):
    for cell in row:
        if isinstance(cell.value, (int, float)):
             cell.number_format = NUM_FORMAT

wb.save(OUTPUT_ANALYTICAL)
print("Visual Precision Overhaul Complete. Charts and Tables Aligned.")
