"""
Unified Sales Report Builder (v3)
---------------------------------
Features:
- Processes "Old Format" (2000-2004) with 15-column layout.
- Processes "Intermediate Format" (IVA Cater 2004-2005) with 20-column layout.
- Processes "New Format" (2011-2021) with 20-column layout.
- Formal Validation:
    * Yellow flag: IVA rate deviation > 1%.
    * Orange flag: Total != sum(components) within $0.02.
    * Traceability: "Ref. Origen" for all flagged rows.
- Sorting:
    1. Group by Year.
    2. Sort chronologically (Date).
    3. Type Priority (FAC before N/C).
"""
import pandas as pd
import numpy as np
import re
import sys
import os
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

# -------------------------------------------------------------------
# Configuration
# -------------------------------------------------------------------
INPUT_DIR = "data/input"
OUTPUT_FILE = r"data\output\All_Sales_Report.xlsx"

IVA_RATES = {0.21: "21%", 0.105: "10.5%", 0.27: "27%"}
RATE_TOLERANCE = 0.01
TOTAL_TOLERANCE = 0.02

YELLOW_FILL = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
ORANGE_FILL = PatternFill(start_color="FFCC99", end_color="FFCC99", fill_type="solid")
RED_FONT = Font(color="FF0000")
ACCOUNTING_FORMAT = '#,##0.00'

NEW_COLUMNS = [
    "Nro", "Fecha", "Tipo", "Comprobante", "Comprobante_dup", "CODIGO",
    "Razon Social", "Cuit / DNI", "Condic.",
    "A_Neto_21", "A_Neto_10_5", "B_Neto_21", "B_Neto_10_5", "IVA_Exento",
    "A_IVA_21", "A_IVA_10_5", "B_IVA_21", "B_IVA_10_5",
    "Reten_Percep", "Total",
]

COMPONENT_COLS = [
    "A_Neto_21", "A_Neto_10_5", "B_Neto_21", "B_Neto_10_5", "IVA_Exento",
    "A_IVA_21", "A_IVA_10_5", "B_IVA_21", "B_IVA_10_5", "Reten_Percep",
]

# -------------------------------------------------------------------
# Helper Functions
# -------------------------------------------------------------------

def safe_float(val):
    if pd.isna(val): return 0.0
    try:
        f = float(val)
        return 0.0 if np.isnan(f) else f
    except: return 0.0

def clean_cuit(cuit_val):
    if pd.isna(cuit_val): return ""
    return str(cuit_val).replace("-", "").replace(".", "").replace(",", "").strip()

def format_comprobante(row_data, tipo_col_idx, num_col_idx, default_letter="A"):
    # Standardize result: [Letter]0001[8-digits]
    tipo_raw = str(row_data[tipo_col_idx]).strip().upper() if pd.notna(row_data[tipo_col_idx]) else ""
    
    # Priority: explicitly look for B or C as single letters or prefixes
    # If not found, default to A (standard behavior)
    letter = default_letter
    if "B" in tipo_raw.split() or tipo_raw.startswith("B"): letter = "B"
    elif "C" in tipo_raw.split() or tipo_raw.startswith("C"): letter = "C"
    elif "A" in tipo_raw.split() or tipo_raw.startswith("A"): letter = "A"
    else: 
        # Fallback for "FAC" or "N/C" without explicit letter -> usually A
        if "FAC" in tipo_raw or "N/C" in tipo_raw or "NDC" in tipo_raw:
            letter = "A"
    
    num_raw = str(row_data[num_col_idx]).strip() if pd.notna(row_data[num_col_idx]) else ""
    if not num_raw: return ""
    
    # Extract digits
    digits = "".join(re.findall(r'\d+', num_raw))
    if not digits: return num_raw
    
    # Standardize to 12 digits (4 for point of sale + 8 for number)
    # If we have More than 12 digits, we take the last 12.
    # We use zfill for string padding to avoid ValueError: Unknown format code 'd' for str
    full_num = digits[-12:].zfill(12)
    return f"{letter}{full_num}"

def is_anulado(razon_social):
    if pd.isna(razon_social): return False
    rs = str(razon_social).upper()
    return "ANULADO" in rs or "COMPROBANTE ANULADO" in rs

# -------------------------------------------------------------------
# Format Loaders
# -------------------------------------------------------------------

def load_old_format(file_path):
    """Processes 15-column layout."""
    basename = os.path.basename(file_path)
    xls = pd.ExcelFile(file_path)
    rows = []
    for sheet in xls.sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet, header=None)
        if len(df) < 7: continue
        data = df.iloc[6:].copy()
        for idx, row in data.iterrows():
            if pd.isna(row[1]) and pd.isna(row[14]): continue
            total = safe_float(row[14])
            if abs(total) < 0.01 or is_anulado(row[5]) or str(row[5]).upper() == "TOTALES": continue
            
            # Hybrid NC/NDC detection (Rules: Check col 2 and col 3)
            tipo_marker = (str(row[2]) + " " + str(row[3])).upper()
            tipo = "N/C" if "NC" in tipo_marker or "NDC" in tipo_marker else "FAC"
            
            comp = format_comprobante(row, 2, 4)
            new_row = {
                "Fecha": pd.to_datetime(row[1], errors='coerce'), "Tipo": tipo, "Comprobante": comp, "Comprobante_dup": comp,
                "Razon Social": row[5], "Cuit / DNI": clean_cuit(row[6]), "Condic.": row[7],
                "IVA_Exento": safe_float(row[10]), "Reten_Percep": safe_float(row[13]), "Total": total,
                "A_Neto_21": 0.0, "A_Neto_10_5": 0.0, "B_Neto_21": 0.0, "B_Neto_10_5": 0.0,
                "A_IVA_21": 0.0, "A_IVA_10_5": 0.0, "B_IVA_21": 0.0, "B_IVA_10_5": 0.0,
                "_src_file": basename, "_src_sheet": sheet, "_src_row": idx + 1
            }
            neto_a, neto_byz, iva_gen = safe_float(row[9]), safe_float(row[8]), safe_float(row[11])
            if neto_a != 0: new_row["A_Neto_21"], new_row["A_IVA_21"] = neto_a, iva_gen
            elif neto_byz != 0: new_row["A_Neto_10_5"], new_row["A_IVA_10_5"] = neto_byz, iva_gen
            rows.append(new_row)
    return rows

def load_type_d_format(file_path):
    """Processes 18-column layout."""
    basename = os.path.basename(file_path)
    xls = pd.ExcelFile(file_path)
    rows = []
    for sheet in xls.sheet_names:
        df = pd.read_excel(file_path, sheet_name=sheet, header=None)
        if len(df) < 7: continue
        data = df.iloc[6:].copy()
        for idx, row in data.iterrows():
            if pd.isna(row[1]) or pd.isna(row[4]): continue
            total = safe_float(row[17])
            if abs(total) < 0.01 or is_anulado(row[4]) or str(row[4]).upper() == "TOTALES":
                continue
            tipo = "N/C" if "N/C" in str(row[2]).upper() else "FAC"
            comp = format_comprobante(row, 2, 3)
            new_row = {
                "Fecha": pd.to_datetime(row[1], errors='coerce'), "Tipo": tipo, "Comprobante": comp, "Comprobante_dup": comp,
                "Razon Social": row[4], "Cuit / DNI": clean_cuit(row[5]), "Condic.": row[6],
                "A_Neto_21": safe_float(row[7]), "A_Neto_10_5": safe_float(row[8]), "B_Neto_21": safe_float(row[9]),
                "B_Neto_10_5": safe_float(row[10]), "IVA_Exento": safe_float(row[11]), "A_IVA_21": safe_float(row[12]),
                "A_IVA_10_5": safe_float(row[13]), "B_IVA_21": safe_float(row[14]), "B_IVA_10_5": safe_float(row[15]),
                "Reten_Percep": safe_float(row[16]), "Total": total,
                "_src_file": basename, "_src_sheet": sheet, "_src_row": idx + 1
            }
            rows.append(new_row)
    return rows

def load_new_format(file_path):
    """Processes 20, 26, or 68+ column layout."""
    basename = os.path.basename(file_path)
    xls = pd.ExcelFile(file_path)
    # Target "ventas nuevas" for intermediate files, else first sheet
    sheet_name = "ventas nuevas" if "ventas nuevas" in xls.sheet_names else 0
    df = pd.read_excel(file_path, sheet_name=sheet_name, header=None)
    if len(df) < 7: return []
    data = df.iloc[6:].copy()
    num_cols = data.shape[1]
    rows = []
    
    # Mappings
    if num_cols == 26: # Type B (IVA Cater 2004-2010)
        idx_rs, idx_total, idx_percep = 5, 18, 17
        idx_n21, idx_n105, idx_bn21, idx_bn105, idx_ex = 8, 9, 10, 11, 12
        idx_i21, idx_i105, idx_bi21, idx_bi105 = 13, 14, 15, 16
        idx_cuit, idx_cond = 7, 6
        idx_codigo = 4
    else: # Type C mapping (20, 68, 69 cols etc. - 2011-2026)
        idx_rs, idx_total, idx_percep = 6, 19, 18
        idx_n21, idx_n105, idx_bn21, idx_bn105, idx_ex = 9, 10, 11, 12, 13
        idx_i21, idx_i105, idx_bi21, idx_bi105 = 14, 15, 16, 17
        idx_cuit, idx_cond = 7, 8
        idx_codigo = 5

    for idx, row in data.iterrows():
        total = safe_float(row[idx_total])
        rs = row[idx_rs]
        if pd.isna(row[1]) or pd.isna(rs): continue
        if is_anulado(rs) or str(rs).upper() == "TOTALES" or abs(total) < 0.01: continue
        tipo = "N/C" if "N/C" in str(row[2]).upper() else "FAC"
        # Standardize Comprobante formatting for monthly files
        comp = format_comprobante(row, 2, 3) if num_cols < 26 else format_comprobante(row, 2, 3, default_letter=str(row[2])[0])
        new_row = {
            "Fecha": pd.to_datetime(row[1], errors='coerce'), "Tipo": tipo, "Comprobante": comp, "Comprobante_dup": comp,
            "CODIGO": row[idx_codigo], "Razon Social": rs,
            "Cuit / DNI": clean_cuit(row[idx_cuit]), "Condic.": row[idx_cond],
            "A_Neto_21": safe_float(row[idx_n21]), "A_Neto_10_5": safe_float(row[idx_n105]),
            "B_Neto_21": safe_float(row[idx_bn21]), "B_Neto_10_5": safe_float(row[idx_bn105]),
            "IVA_Exento": safe_float(row[idx_ex]), "A_IVA_21": safe_float(row[idx_i21]),
            "A_IVA_10_5": safe_float(row[idx_i105]), "B_IVA_21": safe_float(row[idx_bi21]),
            "B_IVA_10_5": safe_float(row[idx_bi105]), "Reten_Percep": safe_float(row[idx_percep]), "Total": total,
            "_src_file": basename, "_src_sheet": sheet_name, "_src_row": idx + 1
        }
        rows.append(new_row)
    return rows

# -------------------------------------------------------------------
# Process
# -------------------------------------------------------------------

def main():
    print("Step 1: Discovering and Loading files...")
    all_records = []
    processed_keys = set()
    
    files = [f for f in os.listdir(INPUT_DIR) if f.endswith(('.xls', '.xlsx'))]
    
    for filename in sorted(files):
        path = os.path.join(INPUT_DIR, filename)
        # Determine format by column count of first data sheet
        try:
            temp_df = pd.read_excel(path, sheet_name=0, header=None, nrows=10)
            cols = temp_df.shape[1]
            if cols == 15: loader = load_old_format
            elif cols == 18: loader = load_type_d_format
            elif cols in (20, 26): loader = load_new_format
            else:
                xls = pd.ExcelFile(path)
                if "ventas nuevas" in xls.sheet_names: loader = load_new_format
                else:
                    print(f"  ? Skipping unknown format: {filename} ({cols} cols)")
                    continue
            
            print(f"  - Loading {filename}...")
            records = loader(path)
            
            # Deduplication logic
            added = 0
            for r in records:
                # Key: (Date, Type, Comprobante, Total rounded)
                key = (r["Fecha"], r["Tipo"], r["Comprobante"], round(r["Total"], 2))
                if key not in processed_keys:
                    all_records.append(r)
                    processed_keys.add(key)
                    added += 1
            if added < len(records):
                print(f"    (Skipped {len(records)-added} duplicates)")
                
        except Exception as e:
            print(f"  [ERROR] Skipping {filename} due to internal error: {e}")
            continue # Resilient: continue with next file instead of crashing

    df = pd.DataFrame(all_records)
    print(f"Step 2: Sorting {len(df)} unique records...")
    df['Year'] = df['Fecha'].dt.year.fillna(0).astype(int)
    df['TipoPriority'] = df['Tipo'].apply(lambda x: 1 if x == 'FAC' else 2)
    df = df.sort_values(by=['Year', 'Fecha', 'TipoPriority', 'Comprobante']).reset_index(drop=True)
    df['Nro'] = df.index + 1
    
    print("Step 2.1: Applying negative signs for N/C entries...")
    # Numeric columns to negate
    numeric_cols = COMPONENT_COLS + ["Total"]
    df.loc[df['Tipo'] == 'N/C', numeric_cols] *= -1

    print("Step 3: Validating and flagging...")
    yellow_cells, orange_cells = [], []
    df["Ref. Origen"] = ""
    for idx, row in df.iterrows():
        flagged = False
        comp_sum = sum(row[c] for c in COMPONENT_COLS)
        if abs(row["Total"] - comp_sum) > TOTAL_TOLERANCE:
            orange_cells.append((idx, "Total", "Mismatch"))
            flagged = True
        for n, i, r in [("A_Neto_21", "A_IVA_21", 0.21), ("A_Neto_10_5", "A_IVA_10_5", 0.105)]:
            if row[n] != 0 and abs(abs(row[i]/row[n]) - r) > RATE_TOLERANCE:
                yellow_cells.append((idx, i, "Rate"))
                flagged = True
        total_neto = row["A_Neto_21"] + row["A_Neto_10_5"] + row["B_Neto_21"] + row["B_Neto_10_5"]
        if abs(total_neto) > 0.1 and abs(row["Reten_Percep"] / total_neto) > 0.10:
            yellow_cells.append((idx, "Reten_Percep", "HighP"))
            flagged = True
        if flagged:
            df.at[idx, "Ref. Origen"] = f"{row['_src_file']} | {row['_src_sheet']} | Row {row['_src_row']}"

    print(f"Step 4: Saving to {OUTPUT_FILE}...")
    final_cols = NEW_COLUMNS + ["Ref. Origen"]
    df[final_cols].to_excel(OUTPUT_FILE, index=False, sheet_name="ventas")

    print("Step 5: Styling (Highlights, Red Font, Number Format)...")
    wb = load_workbook(OUTPUT_FILE); ws = wb["ventas"]
    col_map = {name: i + 1 for i, name in enumerate(final_cols)}
    numeric_indices = [col_map[c] for c in numeric_cols]
    
    # 1. Apply Yellow/Orange Highlights
    for r, c, _ in yellow_cells: ws.cell(row=r+2, column=col_map[c]).fill = YELLOW_FILL
    for r, c, _ in orange_cells: ws.cell(row=r+2, column=col_map[c]).fill = ORANGE_FILL
    
    # 2. Iterative pass for per-row styles (Red Font for N/C and Number Format for all)
    for row_idx in range(2, ws.max_row + 1):
        tipo_val = ws.cell(row=row_idx, column=col_map["Tipo"]).value
        is_nc = (tipo_val == "N/C")
        
        # Apply style to all columns in the row
        for col_idx in range(1, len(final_cols) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            
            # Red Font for N/C entries
            if is_nc:
                cell.font = RED_FONT
            
            # Number Format for numeric columns
            if col_idx in numeric_indices:
                cell.number_format = ACCOUNTING_FORMAT

    wb.save(OUTPUT_FILE)

    print(f"\nSUMMARY\n-------\nTotal Rows: {len(df)}\nYellow: {len(yellow_cells)}\nOrange: {len(orange_cells)}\nDone.")

if __name__ == "__main__":
    main()
