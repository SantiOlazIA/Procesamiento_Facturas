import pandas as pd
import os
import argparse
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
from google.oauth2.service_account import Credentials
from googleapiclient.discovery import build
from googleapiclient.http import MediaFileUpload

DRIVE_FOLDER_ID = '1kuJ7BHR6zXmT3fTZMHQgBHg6yhdGqgG7'
SCOPES = ['https://www.googleapis.com/auth/drive.file']
CREDENTIALS_FILE = r'C:\Users\Tuchi\MiEstudioIA\credenciales_drive.json'

def upload_to_drive(file_path):
    print("\nIniciando subida a Google Drive...")
    try:
        if not os.path.exists(CREDENTIALS_FILE):
            print(f"Error: No se encontró el archivo de credenciales en {CREDENTIALS_FILE}")
            return

        creds = Credentials.from_service_account_file(CREDENTIALS_FILE, scopes=SCOPES)
        service = build('drive', 'v3', credentials=creds)
        file_metadata = {
            'name': os.path.basename(file_path),
            'parents': [DRIVE_FOLDER_ID]
        }
        media = MediaFileUpload(file_path, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        file = service.files().create(body=file_metadata, media_body=media, fields='id').execute()
        print(f"¡Éxito! Archivo subido a Drive correctamente. ID del archivo: {file.get('id')}")

    except Exception:
        # Catch errors without crashing or showing massive trace
        print(f"\n[ADVERTENCIA] No se pudo subir el archivo a Google Drive (Acceso denegado o credenciales inválidas).")
        print("El proceso local finalizó correctamente, puede encontrar el archivo en la carpeta Output.")
def transform_excel(input_file, output_file):
    if not os.path.exists(input_file):
        print(f"Error: No se encuentra el archivo en {input_file}")
        return

    print(f"Procesando archivo: {input_file}")
    try:
        df_orig = pd.read_excel(input_file)
    except Exception as e:
        print(f"Error crítico al leer el archivo Excel: {e}")
        return

    # --- VALIDACIÓN DE COLUMNAS ---
    required_columns = [
        'IDENTIFTRI', 'N_COMP', 'FECHA_EMI', 'PORC_IVA', 'IMP_IVA', 
        'IMP_TOTAL', 'IMP_NETO', 'IMP_EXENTO', 'OTROSIMP', 
        'T_COMP', 'NOM_PROVE', 'COND_IVA'
    ]
    
    missing_cols = [col for col in required_columns if col not in df_orig.columns]
    if missing_cols:
        print(f"\n{'!'*60}")
        print(f"ERROR: El archivo de entrada no tiene el formato esperado.")
        print(f"Faltan las siguientes columnas obligatorias:")
        for col in missing_cols:
            print(f" - {col}")
        print(f"Por favor verifique que el archivo de entrada sea correcto.")
        print(f"{'!'*60}\n")
        return
    # ------------------------------
    columns_target = [
        'Fecha', 'Tipo', 'Comprobante', 'Razón Social', 'Cuit', 'Condic.',
        'Neto Gravado 21%', 'Neto Gravado 10,5%', 'Neto Gravado 27%',
        'Conceptos no Gravados', 'I.V.A 21%', 'I.V.A 10,5%', 'I.V.A 27%',
        'Compras sin IVA', 'Percepción IB San Juan', 'Retenc. / Percepc.', 'Total'
    ]

    temp_rows = []
    if 'IDENTIFTRI' in df_orig.columns and 'N_COMP' in df_orig.columns and 'FECHA_EMI' in df_orig.columns:
        df_orig = df_orig.sort_values(by=['IDENTIFTRI', 'N_COMP', 'FECHA_EMI']).reset_index(drop=True)

    for i, row in df_orig.iterrows():
        n_comp = str(row['N_COMP']).strip()
        p_iva = row['PORC_IVA'] if not pd.isna(row['PORC_IVA']) else 0.0
        imp_iva = row['IMP_IVA']
        imp_total = row['IMP_TOTAL']
        imp_neto = row['IMP_NETO'] if pd.notna(row['IMP_NETO']) else 0.0
        imp_exento = row['IMP_EXENTO'] if pd.notna(row['IMP_EXENTO']) else 0.0
        otros_imp = row['OTROSIMP'] if 'OTROSIMP' in row and pd.notna(row['OTROSIMP']) else 0.0

        if i > 0 and p_iva in [3.0, 1.5] and n_comp == str(df_orig.loc[i-1, 'N_COMP']).strip():
            if temp_rows:
                temp_rows[-1]['Retenc. / Percepc.'] += imp_iva
                temp_rows[-1]['original_total_ref'] += imp_total 
                continue
        
        new_row = {col: 0.0 for col in columns_target}
        new_row['Fecha'] = row['FECHA_EMI']
        new_row['Tipo'] = row['T_COMP']
        new_row['Comprobante'] = n_comp
        new_row['Razón Social'] = row['NOM_PROVE']
        new_row['Cuit'] = row['IDENTIFTRI']
        new_row['Condic.'] = row['COND_IVA']
        new_row['original_total_ref'] = imp_total
        new_row['original_iva_ref'] = imp_iva if p_iva in [21.0, 10.5, 27.0] else 0.0
        
        if p_iva == 21.0:
            new_row['Neto Gravado 21%'] = imp_neto
            new_row['I.V.A 21%'] = imp_iva
        elif p_iva == 10.5:
            new_row['Neto Gravado 10,5%'] = imp_neto
            new_row['I.V.A 10,5%'] = imp_iva
        elif p_iva == 27.0:
            new_row['Neto Gravado 27%'] = imp_neto
            new_row['I.V.A 27%'] = imp_iva
        elif p_iva in [3.0, 1.5]: 
            new_row['Retenc. / Percepc.'] = imp_iva

        neto_huerfano = imp_neto if p_iva not in [21.0, 10.5, 27.0, 3.0, 1.5] else 0.0
        new_row['Percepción IB San Juan'] = otros_imp

        is_factura_c = n_comp.upper().startswith('C')
        if is_factura_c:
            new_row['Compras sin IVA'] = imp_exento + neto_huerfano
        else:
            new_row['Conceptos no Gravados'] = imp_exento + neto_huerfano
            
        temp_rows.append(new_row)

    df_intermediate = pd.DataFrame(temp_rows)
    group_cols = ['Fecha', 'Tipo', 'Comprobante', 'Razón Social', 'Cuit', 'Condic.']
    
    numeric_cols = [col for col in columns_target if col not in group_cols] + ['original_total_ref', 'original_iva_ref']
    df_consolidated = df_intermediate.groupby(group_cols, as_index=False)[numeric_cols].sum()

    def adjust_otros_imp(row):
        otros_imp = row['Percepción IB San Juan']
        if otros_imp == 0:
            return otros_imp, row['Conceptos no Gravados']
        
        base_imponible = row['Neto Gravado 21%'] + row['Neto Gravado 10,5%'] + row['Neto Gravado 27%'] + row['Conceptos no Gravados'] + row['Compras sin IVA']
        perc_3 = base_imponible * 0.03
        perc_1 = base_imponible * 0.01
        
        if abs(otros_imp - perc_3) <= 0.10:
            return otros_imp, row['Conceptos no Gravados']
        else:
            saldo = otros_imp - perc_1
            return perc_1, row['Conceptos no Gravados'] + saldo

    ajustes = df_consolidated.apply(lambda row: pd.Series(adjust_otros_imp(row)), axis=1)
    df_consolidated['Percepción IB San Juan'] = ajustes[0]
    df_consolidated['Conceptos no Gravados'] = ajustes[1]

    cols_a_sumar = columns_target[6:16]
    df_consolidated['Total'] = df_consolidated[cols_a_sumar].sum(axis=1)

    def check_diff(row):
        total_diff = abs(row['Total'] - row['original_total_ref']) > 0.02
        iva_calc = row['I.V.A 21%'] + row['I.V.A 10,5%'] + row['I.V.A 27%']
        iva_diff = abs(iva_calc - row['original_iva_ref']) > 0.02
        return total_diff or iva_diff

    df_consolidated['ERROR_VALIDACION'] = df_consolidated.apply(check_diff, axis=1)

    # ========================================================
    # 6. ORDENAMIENTO FINAL 
    # ========================================================
    def get_sort_priority(row):
        tipo = str(row['Tipo']).upper()
        comp = str(row['Comprobante']).upper()
        
        # 3 = Notas de Crédito (Van al final)
        if 'N/C' in tipo or 'NC' in tipo or 'CREDITO' in tipo:
            return 3
        # 1 = Facturas A (Van primero)
        elif comp.startswith('A'):
            return 1
        # 2 = Facturas C (Van segundo)
        elif comp.startswith('C'):
            return 2
        # 1.5 = Cualquier otro (Facturas B, M, tickets)
        else:
            return 1.5

    df_consolidated['Prioridad_Orden'] = df_consolidated.apply(get_sort_priority, axis=1)
    
    # Ordenar por: 1° Categoría, 2° Razón Social, 3° Comprobante
    df_consolidated = df_consolidated.sort_values(
        by=['Prioridad_Orden', 'Razón Social', 'Comprobante'],
        ascending=[True, True, True]
    ).reset_index(drop=True)

    # Limpiar columna temporal de prioridad
    df_final = df_consolidated.drop(columns=['Prioridad_Orden'])
    
    # ========================================================
    # 7. GUARDADO Y FORMATO
    # ========================================================
    df_final_to_save = df_final[columns_target + ['ERROR_VALIDACION']]
    os.makedirs(os.path.dirname(output_file), exist_ok=True)
    df_final_to_save.to_excel(output_file, index=False)

    wb = load_workbook(output_file)
    ws = wb.active
    yellow_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")
    
    error_col_index = len(columns_target) + 1 
    
    for row_idx in range(2, ws.max_row + 1):
        if ws.cell(row=row_idx, column=error_col_index).value == True:
            for col_idx in range(1, len(columns_target) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = yellow_fill

    ws.delete_cols(error_col_index)
    wb.save(output_file)
    print(f"--- PROCESO LOCAL FINALIZADO ---")
    print(f"Archivo generado en: {output_file}")

    upload_to_drive(output_file)

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Transformador de Excel de Compras Tango.")
    parser.add_argument("--input", default=r'C:\Users\Tuchi\MiEstudioIA\Input\202601 - IVA Compras.xlsx', help="Ruta del archivo original")
    parser.add_argument("--output", default=r'C:\Users\Tuchi\MiEstudioIA\Output\Gama_Compras_Procesado_Final.xlsx', help="Ruta de guardado")
    args = parser.parse_args()

    transform_excel(args.input, args.output)