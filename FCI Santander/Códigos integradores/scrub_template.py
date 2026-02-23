"""Scrub the template file to remove all data and formatting below row 4."""
import openpyxl
from openpyxl.styles import PatternFill, Border, Font, Alignment
import os
import shutil

TEMPLATE_PATH = r"C:\Users\Tuchi\MiEstudioIA\FCI Santander\Input\Formato FCI v2.xlsx"
BACKUP_PATH = TEMPLATE_PATH + ".backup"

# Create a backup just in case
if not os.path.exists(BACKUP_PATH):
    shutil.copy2(TEMPLATE_PATH, BACKUP_PATH)

wb = openpyxl.load_workbook(TEMPLATE_PATH)
ws = wb.active

# Clear everything below row 4 for columns A to AZ
for r in range(5, max(ws.max_row + 1, 1500)):
    for col_idx in range(1, 40):
        cell = ws.cell(row=r, column=col_idx)
        cell.value = None
        # Remove all styles
        cell.fill = PatternFill(fill_type=None)
        cell.border = Border()
        cell.font = Font()

# Save the scrubbed template
wb.save(TEMPLATE_PATH)
print("Template scrubbed successfully!")
