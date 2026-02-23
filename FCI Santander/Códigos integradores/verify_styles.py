"""Verify formatting of the generated generated Cuadro rows."""
import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Tuchi\MiEstudioIA\FCI Santander\Output\FCI_Procesado_FCI_SUPER_AHORRO_CL_H_PJ.xlsx")
ws = wb.active

print("=== Formatting of generated row 210 ===")
for c in ['V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG']:
    cell = ws[f'{c}210']
    font = cell.font
    align = cell.alignment
    border = cell.border
    
    bt = border.top.style if border.top else None
    bb = border.bottom.style if border.bottom else None
    bl = border.left.style if border.left else None
    br = border.right.style if border.right else None
    
    print(f"Col {c}: Font={font.name} {font.size}, Align={align.horizontal}, "
          f"Borders(T/B/L/R)={bt}/{bb}/{bl}/{br}")

print("\n=== And row 216 (last row) to check bottom border ===")
for c in ['V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG']:
    cell = ws[f'{c}216']
    border = cell.border
    bb = border.bottom.style if border.bottom else None
    print(f"Col {c}: Bottom Border={bb}")
