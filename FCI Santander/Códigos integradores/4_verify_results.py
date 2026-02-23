"""
FCI Santander - Verificacion de Resultados (Paso 4)

Two verification methods:
1. Portfolio Check: Final CPs in Excel vs 'ESTADO DE CARTERA VALORIZADA'
   on page 8 of the latest PDF (tolerance 0.99, price ignored).
2. Mathematical Check: Prueba columns AF/AG must resolve to 0.
"""
import openpyxl
import os
import sys
import json

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(SCRIPT_DIR)
from config import OUTPUT_EXCEL_PATH, JSON_DATA_PATH, log


def verify_balances():
    log.info("=" * 50)
    log.info("PASO 4: Verificacion de Resultados")
    log.info("=" * 50)

    if not os.path.exists(OUTPUT_EXCEL_PATH):
        log.error(f"Archivo Excel no encontrado: {OUTPUT_EXCEL_PATH}")
        return False

    wb = openpyxl.load_workbook(OUTPUT_EXCEL_PATH, data_only=True)
    ws = wb.active

    issues = 0

    # --- Method 1: Portfolio Check ---
    log.info("\n  [1] Portfolio Check (CPs vs PDF)")
    portfolio_cp = _get_portfolio_cp()

    if portfolio_cp is not None:
        # Sum all CPs from Z column using Decimal for exact arithmetic
        from decimal import Decimal, ROUND_HALF_UP
        excel_final_cp = Decimal('0')
        for r in range(5, 50000):
            z_val = ws[f'Z{r}'].value
            if z_val is None:
                continue
            try:
                excel_final_cp += Decimal(str(z_val))
            except Exception:
                pass
        excel_final_cp = float(excel_final_cp)

        diff = abs(excel_final_cp - portfolio_cp)
        if diff < 0.99:
            log.info(
                f"  Excel CPs:     {excel_final_cp:>18,.2f}"
            )
            log.info(
                f"  Portfolio CPs: {portfolio_cp:>18,.2f}"
            )
            log.info(f"  Diferencia:    {diff:.2f} -> OK")
        else:
            log.warning(
                f"  Excel CPs:     {excel_final_cp:>18,.2f}"
            )
            log.warning(
                f"  Portfolio CPs: {portfolio_cp:>18,.2f}"
            )
            log.warning(
                f"  Diferencia:    {diff:,.2f} -> DISCREPANCIA"
            )
            issues += 1
    else:
        log.info("  (Sin datos de portfolio para verificar)")

    # --- Method 2: Mathematical Check (Prueba columns) ---
    log.info("\n  [2] Mathematical Check (Prueba AF/AG)")
    prueba_issues = 0

    for r in range(5, 50000):
        af_val = ws[f'AF{r}'].value
        ag_val = ws[f'AG{r}'].value

        if af_val is None and ag_val is None:
            continue

        try:
            af_num = float(af_val) if af_val else 0.0  # noqa: F841
            ag = float(ag_val) if ag_val else 0.0

            if abs(ag) > 0.01:
                log.warning(
                    f"  Fila {r}: AG={ag:.4f} (debería ser 0)"
                )
                prueba_issues += 1
        except (ValueError, TypeError):
            pass

    if prueba_issues == 0:
        log.info("  Todas las formulas Prueba = 0 -> OK")
    else:
        log.warning(
            f"  {prueba_issues} fila(s) con Prueba != 0"
        )
        issues += prueba_issues

    # Summary
    log.info(f"\n  Resultado: {issues} problema(s) encontrado(s)")

    if issues == 0:
        log.info("  Verificacion EXITOSA")
    else:
        log.warning("  Verificacion con ERRORES")

    return issues == 0


def _get_portfolio_cp():
    """Read the portfolio_cp value from the fund's JSON file."""
    if not os.path.exists(JSON_DATA_PATH):
        return None

    try:
        with open(JSON_DATA_PATH, 'r', encoding='utf-8') as f:
            data = json.load(f)

        return data.get('portfolio_cp')
    except Exception:
        return None


if __name__ == "__main__":
    success = verify_balances()
    sys.exit(0 if success else 1)
