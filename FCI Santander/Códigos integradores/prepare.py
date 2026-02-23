"""
FCI Santander - Prepare Environment (Paso 2.5)

Cleans the output template and ensures a fresh start for the fund.
NO hardcoded balances (initial balance handled by Step 3).
"""
import openpyxl
from openpyxl.styles import PatternFill, Border
import os
import sys
import shutil

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.append(SCRIPT_DIR)
from config import OUTPUT_EXCEL_PATH, TEMPLATE_EXCEL_PATH, log

# Template source from Input folder
TEMPLATE_SRC = os.path.join(os.path.dirname(SCRIPT_DIR), 'Input', 'Formato FCI v2.xlsx')


def prepare():
    """
    Copy and clean the template for the current fund.
    """
    log.info("=" * 50)
    log.info("PREPARACION DEL ENTORNO")
    log.info("=" * 50)

    if os.path.exists(OUTPUT_EXCEL_PATH):
        try:
            os.remove(OUTPUT_EXCEL_PATH)
            log.info("Eliminado output anterior")
        except Exception as e:
            log.warning(f"No se pudo eliminar {OUTPUT_EXCEL_PATH}: {e}")

    # Use Template_FCI.xlsx if it exists, otherwise fall back to the source Input
    if os.path.exists(TEMPLATE_EXCEL_PATH):
        source = TEMPLATE_EXCEL_PATH
    elif os.path.exists(TEMPLATE_SRC):
        source = TEMPLATE_SRC
        # Ensure it exists in the integradores folder too for next time
        shutil.copy2(TEMPLATE_SRC, TEMPLATE_EXCEL_PATH)
    else:
        log.error(f"Template no encontrado en {TEMPLATE_SRC} ni {TEMPLATE_EXCEL_PATH}")
        return False

    shutil.copy2(source, OUTPUT_EXCEL_PATH)
    wb = openpyxl.load_workbook(OUTPUT_EXCEL_PATH)
    ws = wb.active

    # Clear table data (W-AG, rows 5+)
    # Note: Row 4 is the header
    log.debug("Limpiando datos del cuadro (filas 5+)...")
    for r in range(5, 10000):
        has_data = False
        for c in ['V', 'W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG']:
            if ws[f'{c}{r}'].value is not None:
                ws[f'{c}{r}'].value = None
                ws[f'{c}{r}'].fill = PatternFill(fill_type=None)
                has_data = True
        if not has_data and r > 100:  # Optimization
            break

    # Clear asiento data (B-U, rows 4+)
    log.debug("Limpiando asientos contables...")
    for r in range(4, 10000):
        has_data = False
        for c in ['B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U']:
            if ws[f'{c}{r}'].value is not None:
                ws[f'{c}{r}'].value = None
                has_data = True
        # Clear borders in the accounting block
        for c in ['B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{c}{r}'].border = Border()
        
    # Update column headers
    ws['Z4'] = "CUOTA PARTE"
    ws['AA4'] = "VALOR CUOTA PARTE"
    ws['AB4'] = "IMPORTE CTA CTE"

    # Update branding headers
    fund_name = os.environ.get('FCI_CURRENT_FUND', 'Santander')
    ws['E1'] = 'CaterWest'
    ws['B1'] = 'Ejercicio 2025/26'
    ws['B3'] = 'ASIENTOS INVERSIONES FINANCIERAS Santander'
    ws['W3'] = f'FONDO COMUN DE INVERSION Santander {fund_name} (primero entrado primero salido)'

    wb.save(OUTPUT_EXCEL_PATH)
    log.info("Template preparado y limpio correctamente")
    return True


if __name__ == "__main__":
    success = prepare()
    sys.exit(0 if success else 1)
