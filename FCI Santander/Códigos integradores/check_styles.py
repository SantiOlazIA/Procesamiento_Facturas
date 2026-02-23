"""Check the exact formatting of the original template for a typical Cuadro row."""
import openpyxl

wb = openpyxl.load_workbook(r"C:\Users\Tuchi\MiEstudioIA\FCI Santander\Input\Formato_FCI_v2_backup.xlsx")
ws = wb.active

# Check row 10 in columns W through AG
print("=== Original Template Formatting (Row 10) ===")
for c in ['W', 'X', 'Y', 'Z', 'AA', 'AB', 'AC', 'AD', 'AE', 'AF', 'AG']:
    cell = ws[f'{c}10']
    font = cell.font
    align = cell.alignment
    border = cell.border
    
    # Borders
    bt = border.top.style if border.top else None
    bb = border.bottom.style if border.bottom else None
    bl = border.left.style if border.left else None
    br = border.right.style if border.right else None
    
    print(f"Col {c}: Font={font.name} {font.size}, Align={align.horizontal}, "
          f"Borders(T/B/L/R)={bt}/{bb}/{bl}/{br}")

print("\n=== And row 10 in Asiento (B-H) ===")
for c in ['B', 'C', 'D', 'G', 'H']:
    cell = ws[f'{c}10']
    font = cell.font
    align = cell.alignment
    
    print(f"Col {c}: Font={font.name} {font.size}, Align={align.horizontal}")
