import openpyxl
import os
import re

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
EXCEL_PATH = os.path.join(SCRIPT_DIR, 'FCI_Procesado_Anual.xlsx')

def check_column_m():
    if not os.path.exists(EXCEL_PATH):
        print("Archivo no encontrado")
        return

    wb = openpyxl.load_workbook(EXCEL_PATH, data_only=False)
    ws = wb.active
    
    # Column M is index 13 (1-based) -> 'M'
    
    formulas = []
    non_zero_formulas = []
    
    last_row = ws.max_row
    print(f"Escaneando Column M hasta fila {last_row}...")
    
    count_m = 0
    
    for r in range(4, last_row + 1):
        cell_m = ws[f'M{r}']
        val_m = cell_m.value
        
        if val_m:
            count_m += 1
            # Check if it looks like =L{row}-H{something}
            # Expected: =L{r}-H{r+3}
            
            # Simple check: Does it subtract two things?
            formulas.append(str(val_m))
            
            # Verify the logic L - H where L points to H
            # L cell:
            cell_l = ws[f'L{r}']
            val_l = cell_l.value # Expected =H{r+3}
            
            # M formula expected: =L{r}-H{r+3}
            expected_m = f"=L{r}-H{r+3}"
            
            if val_m != expected_m:
                print(f"Fila {r}: Formula inesperada en M: {val_m} (Esperaba {expected_m})")
                non_zero_formulas.append((r, val_m))
            else:
                # Check L ref
                expected_l = f"=H{r+3}"
                if val_l != expected_l:
                     print(f"Fila {r}: Formula en M es correcta pero L es sospechosa: {val_l} (Esperaba {expected_l})")
                     non_zero_formulas.append((r, f"M ok but L={val_l}"))

    if not formulas:
        print("No se encontraron valores en la columna M.")
    else:
        print(f"Se encontraron {len(formulas)} celdas con datos en columna M.")
        if not non_zero_formulas:
            print("TODAS las formulas en M siguen el patrón '=L{r}-H{r+3}' donde 'L{r}=H{r+3}'.")
            print("Por lo tanto, el resultado matematico es EXACTAMENTE CERO (0.00).")
        else:
            print("Se encontraron formulas que podrian no dar cero.")

if __name__ == "__main__":
    check_column_m()
