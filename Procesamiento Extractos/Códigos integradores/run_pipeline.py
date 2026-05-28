"""
Pipeline de Extractos Bancarios — Orquestador
Uso: python run_pipeline.py [--period YYYYMM]

Flujo:
  1. Detectar PDFs → extraer transacciones
  2. Clasificar por categoría
  3. Mostrar revisión de clasificación
  4. Confirmar (s/n) → generar Excel
  5. Mostrar resumen de totales
  6. Verificar totales vs. extractos oficiales
"""
import argparse
import os
import sys
from collections import defaultdict
from datetime import date as _date

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
sys.path.insert(0, SCRIPT_DIR)

from config import (log, OUTPUT_DIR, INPUT_DIR, CATEGORIES, CAT_NONE,
                    FONT_NAME, FONT_SIZE, DATE_FORMAT,
                    HEADER_FILL_COLOR, HEADER_FONT_COLOR,
                    TOTAL_FILL_COLOR, ALT_ROW_COLOR, NONE_ROW_COLOR)
from extractor  import extract_all
from classifier import classify_all, build_review_report, build_totals_summary
from verify     import verify_all

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side, numbers


# ─────────────────────────────────────────────────────────────────────────────
# Excel helpers
# ─────────────────────────────────────────────────────────────────────────────

def _hfill(hex_color: str) -> PatternFill:
    return PatternFill('solid', fgColor=hex_color)


def _font(bold=False, color='000000', size=None) -> Font:
    return Font(name=FONT_NAME, bold=bold, color=color, size=size or FONT_SIZE)


def _border_all() -> Border:
    thin = Side(border_style='thin', color='CCCCCC')
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _auto_width(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = col[0].column_letter
        for cell in col:
            try:
                val = str(cell.value) if cell.value is not None else ''
                max_len = max(max_len, len(val))
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = min(max_len + 3, 60)


NUM_FMT = '#.##0,00'   # Argentine style for Excel (Excel reads this fine in ES locale)


def _write_resumen_sheet(wb, all_transactions: dict):
    ws = wb.create_sheet("RESUMEN", 0)

    # Header row
    headers = ['Banco'] + CATEGORIES + ['TOTAL']
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = _hfill(HEADER_FILL_COLOR)
        cell.font = _font(bold=True, color=HEADER_FONT_COLOR)
        cell.alignment = Alignment(horizontal='center', wrap_text=True)
        cell.border = _border_all()
    ws.row_dimensions[1].height = 35

    # Calculate totals
    bank_totals = defaultdict(lambda: defaultdict(float))
    for bank, txns in all_transactions.items():
        for txn in txns:
            if txn['category'] != CAT_NONE:
                bank_totals[bank][txn['category']] += txn['debit']

    grand = defaultdict(float)
    row = 2
    for bank in sorted(bank_totals.keys()):
        ws.cell(row=row, column=1, value=bank).font = _font(bold=True)
        ws.cell(row=row, column=1).border = _border_all()
        row_total = 0.0
        for col_idx, cat in enumerate(CATEGORIES, start=2):
            v = bank_totals[bank].get(cat, 0.0)
            grand[cat] += v
            row_total  += v
            cell = ws.cell(row=row, column=col_idx, value=v if v else None)
            cell.number_format = NUM_FMT
            cell.border = _border_all()
            cell.alignment = Alignment(horizontal='right')
        # Total columna
        cell = ws.cell(row=row, column=len(headers), value=row_total if row_total else None)
        cell.number_format = NUM_FMT
        cell.font = _font(bold=True)
        cell.border = _border_all()
        cell.alignment = Alignment(horizontal='right')
        row += 1

    # TOTAL row
    ws.cell(row=row, column=1, value='TOTAL').font = _font(bold=True)
    ws.cell(row=row, column=1).fill = _hfill(TOTAL_FILL_COLOR)
    ws.cell(row=row, column=1).border = _border_all()
    grand_total = 0.0
    for col_idx, cat in enumerate(CATEGORIES, start=2):
        v = grand[cat]
        grand_total += v
        cell = ws.cell(row=row, column=col_idx, value=v if v else None)
        cell.number_format = NUM_FMT
        cell.fill = _hfill(TOTAL_FILL_COLOR)
        cell.font = _font(bold=True)
        cell.border = _border_all()
        cell.alignment = Alignment(horizontal='right')
    cell = ws.cell(row=row, column=len(headers), value=grand_total)
    cell.number_format = NUM_FMT
    cell.fill = _hfill(TOTAL_FILL_COLOR)
    cell.font = _font(bold=True)
    cell.border = _border_all()
    cell.alignment = Alignment(horizontal='right')

    _auto_width(ws)


def _write_bank_sheet(wb, bank: str, transactions: list[dict]):
    # Nombre de hoja: máx 31 chars, sin caracteres inválidos
    sheet_name = bank[:31].replace('/', '-')
    ws = wb.create_sheet(sheet_name)

    headers = ['Fecha', 'Descripción', 'Débito', 'Crédito', 'Categoría']
    for col_idx, h in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.fill = _hfill(HEADER_FILL_COLOR)
        cell.font = _font(bold=True, color=HEADER_FONT_COLOR)
        cell.alignment = Alignment(horizontal='center')
        cell.border = _border_all()

    # Agrupar subtotales al final
    cat_subtotals = defaultdict(float)

    for row_idx, txn in enumerate(transactions, start=2):
        is_none = txn['category'] == CAT_NONE
        alt     = row_idx % 2 == 0

        fill = _hfill(NONE_ROW_COLOR) if is_none else (_hfill(ALT_ROW_COLOR) if alt else None)

        def wcell(col, value, num_fmt=None, bold=False):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.border = _border_all()
            cell.font   = _font(bold=bold)
            if fill:
                cell.fill = fill
            if num_fmt:
                cell.number_format = num_fmt
                cell.alignment = Alignment(horizontal='right')
            return cell

        date_val = _date.fromisoformat(txn['date']) if txn['date'] else None
        wcell(1, date_val, num_fmt=DATE_FORMAT)
        wcell(2, txn['description'])
        wcell(3, txn['debit']  if txn['debit']  else None, num_fmt=NUM_FMT)
        wcell(4, txn['credit'] if txn['credit'] else None, num_fmt=NUM_FMT)
        wcell(5, txn['category'])

        if txn['debit']:
            cat_subtotals[txn['category']] += txn['debit']

    # Subtotales por categoría
    sub_row = len(transactions) + 3
    ws.cell(row=sub_row - 1, column=1, value='SUBTOTALES POR CATEGORÍA').font = _font(bold=True)
    for cat, total in sorted(cat_subtotals.items()):
        cell_cat = ws.cell(row=sub_row, column=2, value=cat)
        cell_cat.font = _font(bold=True)
        cell_cat.fill = _hfill(TOTAL_FILL_COLOR)
        cell_cat.border = _border_all()
        cell_tot = ws.cell(row=sub_row, column=3, value=total)
        cell_tot.number_format = NUM_FMT
        cell_tot.font = _font(bold=True)
        cell_tot.fill = _hfill(TOTAL_FILL_COLOR)
        cell_tot.border = _border_all()
        cell_tot.alignment = Alignment(horizontal='right')
        sub_row += 1

    _auto_width(ws)


def write_excel(all_transactions: dict, output_path: str):
    wb = openpyxl.Workbook()
    # Remover hoja default
    if 'Sheet' in wb.sheetnames:
        del wb['Sheet']

    _write_resumen_sheet(wb, all_transactions)

    for bank in sorted(all_transactions.keys()):
        txns = all_transactions[bank]
        if txns:
            _write_bank_sheet(wb, bank, txns)

    wb.save(output_path)
    log.info(f"Excel guardado en: {output_path}")


# ─────────────────────────────────────────────────────────────────────────────
# Main
# ─────────────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description='Pipeline de Extractos Bancarios')
    parser.add_argument('--period', type=str, default=None,
                        help='Filtrar por período YYYYMM (ej. 202602). '
                             'Si se omite, procesa todos los PDFs.')
    args = parser.parse_args()

    print()
    print("=" * 68)
    print("  PIPELINE DE EXTRACTOS BANCARIOS — Caterwest SA")
    print("=" * 68)

    # 1. Extraer
    log.info("\nPaso 1: Extrayendo transacciones...")
    all_txns = extract_all(INPUT_DIR, period_filter=args.period)
    if not all_txns:
        print("\n[ERROR] No se encontraron transacciones. Verificar PDFs en:")
        print(f"  {INPUT_DIR}")
        sys.exit(1)

    # 2. Clasificar
    log.info("\nPaso 2: Clasificando...")
    classify_all(all_txns)

    # 3. Revisar
    print()
    print(build_review_report(all_txns))

    # 4. Confirmar
    print()
    try:
        resp = input("Confirmar y generar Excel? (s/n): ").strip().lower()
    except (KeyboardInterrupt, EOFError):
        print("\nAbortado.")
        sys.exit(0)

    if resp not in ('s', 'si', 'sí', 'y', 'yes'):
        print("Cancelado. Podés editar config.py y volver a correr.")
        sys.exit(0)

    # 5. Generar Excel
    period_str = args.period or "multi"
    output_path = os.path.join(OUTPUT_DIR, f"Extractos_{period_str}.xlsx")
    log.info(f"\nPaso 3: Generando Excel → {output_path}")
    write_excel(all_txns, output_path)

    # 6. Resumen final
    print(build_totals_summary(all_txns))
    print(f"\n  Excel guardado en: {output_path}\n")

    # 7. Verificación de totales vs. extractos oficiales
    log.info("Paso 4: Verificando totales contra extractos oficiales...")
    try:
        verify_report = verify_all(all_txns, INPUT_DIR, period_filter=args.period)
        print(verify_report)
    except Exception as e:
        print(f"\n[AVISO] Error durante verificación (no crítico): {e}")


if __name__ == '__main__':
    main()
