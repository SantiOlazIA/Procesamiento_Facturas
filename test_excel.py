import openpyxl

wb = openpyxl.load_workbook(r'C:\Users\Tuchi\MiEstudioIA\FCI Santander\Output\FCI_Procesado_FCI_SUPER_AHORRO_CL_H_PJ.xlsx', data_only=True)
ws = wb.active

for i in range(5, 12):
    date = ws[f'W{i}'].value
    if date:
        cp = ws[f'Z{i}'].value
        price = ws[f'AA{i}'].value
        renta = ws[f'AD{i}'].value
        importe = ws[f'AC{i}'].value
        print(f"Row {i:02d}: {date.strftime('%d/%m/%Y')} | CPs: {cp:14,.2f} | Precio: {price:9.6f} | Importe: {importe:15,.2f} | Renta: {renta:12,.2f}")
